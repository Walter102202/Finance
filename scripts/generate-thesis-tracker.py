# DISCLAIMER: Financial data hardcoded in this script is illustrative only (as of Feb 2026). Update with current data before use. Not investment advice.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import sys, datetime

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
DATE = datetime.date.today().isoformat()
wb = openpyxl.Workbook()

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
blk = Font(name="Arial", color="000000", size=10)
bld = Font(name="Arial", bold=True, size=10)
tf = Font(name="Arial", bold=True, size=14, color="1B3A5C")
sf = Font(name="Arial", bold=True, size=11, color="2C5F8A")
bdr = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"), top=Side("thin","CCCCCC"), bottom=Side("thin","CCCCCC"))
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

def hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bdr

# ========== SHEET 1: Thesis Pillars ==========
ws = wb.active
ws.title = "Thesis"
ws.sheet_properties.tabColor = "1B3A5C"

ws["A1"] = f"Thesis Tracker — {TICKER} (Uber Technologies)"
ws["A1"].font = tf
ws["A2"] = f"Last Updated: {DATE}"
ws["A2"].font = Font(name="Arial", italic=True, color="888888", size=10)

headers = ["#", "Thesis Pillar", "Description", "KPI", "Expected", "Actual", "Status", "Conviction (1-5)", "Last Checked"]
widths = [5, 22, 40, 22, 16, 16, 14, 16, 14]
for i, h in enumerate(headers):
    ws.cell(row=4, column=1+i, value=h)
hdr(ws, 4, len(headers))

pillars = [
    ["1", "Global Mobility Dominance",
     "~75% U.S. ride-hailing share, #1-2 in most international markets. 202M MAPCs, 13.6B trips. Network effects create structural moat.",
     "MAPCs (M)", ">200M (FY2025)", "202M+ (FY2025A)", "On Track", 5, DATE],
    ["2", "AV Platform Orchestration",
     "Multi-partner AV strategy (Waymo, WeRide, Nuro/Lucid, Momenta, Wayve). Positioning as world's largest AV aggregation platform.",
     "AV Cities Live", "10+ by mid-2026", "5 (Austin, Atlanta expanding)", "On Track", 4, DATE],
    ["3", "Advertising Revenue Engine",
     "Uber Ads at $1.5B+ annual run rate, +60% YoY growth. Journey Ads, sponsored listings, merchant offers. ~80% incremental margins.",
     "Ads ARR ($B)", ">$1.5B (FY2025)", "$1.5B+ (Q1 2025 ARR)", "On Track", 5, DATE],
    ["4", "Uber One Membership Flywheel",
     "46M subscribers (+55% YoY). Members spend 3x more, ~50% of total gross bookings. Drives engagement and retention.",
     "Uber One Subs (M)", ">40M (FY2025)", "46M (FY2025A)", "On Track", 5, DATE],
    ["5", "Delivery Margin Expansion",
     "Delivery EBITDA margin expanding rapidly (2.8% of GBs Q4 2023 to 3.6% Q4 2024). Approaching Mobility-like margins long term.",
     "Del. EBITDA % GBs", ">3.5% (Q4 2025)", "~3.6% (Q4 2025A)", "On Track", 4, DATE],
    ["6", "International Growth",
     "60% of Mobility GBs international. Strong MAPCs growth in LatAm, India, Africa, SE Asia. Underpenetrated markets.",
     "Intl Mobility GB Growth", ">15% YoY", "~20% (FY2025A est.)", "On Track", 5, DATE],
    ["7", "FCF & Capital Returns",
     "FCF of $9.8B in FY2025 (+41% YoY). $10B+ buyback capacity. Share count declining. Operating leverage driving expansion.",
     "FCF ($B)", ">$9B (FY2025)", "$9.8B (FY2025A)", "On Track", 5, DATE],
    ["8", "Near-Term Margin Investment",
     "Q1 2026 EBITDA guidance below consensus ($2.37-2.47B vs $2.55B est). Suggests near-term investment in driver supply and intl expansion.",
     "Adj. EBITDA Margin %", ">17% (FY2026E)", "Q1 guide below street", "At Risk", 3, DATE],
]

for i, row in enumerate(pillars):
    r = 5 + i
    for j, val in enumerate(row):
        cell = ws.cell(row=r, column=1+j, value=val)
        cell.font = blue if j in [4, 5] else blk
        cell.alignment = lft if j in [1, 2, 3] else ctr
        cell.border = bdr

# Conditional formatting for Status column (G)
green_fill = PatternFill("solid", fgColor="C8E6C9")
yellow_fill = PatternFill("solid", fgColor="FFF3CD")
red_fill = PatternFill("solid", fgColor="FFCDD2")
ws.conditional_formatting.add("G5:G12", CellIsRule(operator="equal", formula=['"On Track"'], fill=green_fill))
ws.conditional_formatting.add("G5:G12", CellIsRule(operator="equal", formula=['"At Risk"'], fill=yellow_fill))
ws.conditional_formatting.add("G5:G12", CellIsRule(operator="equal", formula=['"Broken"'], fill=red_fill))

# Overall conviction
ws.cell(row=14, column=1, value="Overall Conviction Score").font = sf
ws.cell(row=14, column=8).value = "=AVERAGE(H5:H12)"
ws.cell(row=14, column=8).font = Font(name="Arial", bold=True, size=14, color="1B3A5C")
ws.cell(row=14, column=8).number_format = "0.0"
ws.cell(row=14, column=8).alignment = ctr

ws.cell(row=15, column=1, value="Pillars On Track").font = bld
ws.cell(row=15, column=8).value = '=COUNTIF(G5:G12,"On Track")'
ws.cell(row=15, column=8).font = blk
ws.cell(row=15, column=8).alignment = ctr

ws.cell(row=16, column=1, value="Pillars At Risk").font = bld
ws.cell(row=16, column=8).value = '=COUNTIF(G5:G12,"At Risk")'
ws.cell(row=16, column=8).alignment = ctr

ws.cell(row=17, column=1, value="Pillars Broken").font = bld
ws.cell(row=17, column=8).value = '=COUNTIF(G5:G12,"Broken")'
ws.cell(row=17, column=8).alignment = ctr

for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# ========== SHEET 2: Change Log ==========
ws2 = wb.create_sheet("Change Log")
ws2.sheet_properties.tabColor = "2C5F8A"

ws2["A1"] = "Thesis Change Log"
ws2["A1"].font = tf

log_headers = ["Date", "Trigger / Event", "Pillar Affected", "Old Status", "New Status", "Old Conviction", "New Conviction", "Notes"]
log_widths = [14, 30, 18, 14, 14, 16, 16, 40]
for i, h in enumerate(log_headers):
    ws2.cell(row=3, column=1+i, value=h)
hdr(ws2, 3, len(log_headers))

log_entries = [
    [DATE, "Initiation of coverage", "All", "N/A", "On Track", "N/A", "4-5", "All pillars initialized at coverage initiation."],
    ["2026-02-04", "UBER Q4 2025 Earnings", "Near-Term Margin (#8)", "N/A", "At Risk", "N/A", "3", "Q1 2026 EBITDA guidance $2.37-2.47B below consensus ~$2.55B. Stock fell ~7%. Suggests near-term driver incentive/intl investment."],
    ["2026-02-04", "UBER Q4 2025 Earnings", "Uber One (#4)", "N/A", "On Track", "N/A", "5", "46M subscribers (+55% YoY), now ~50% of GBs. Exceptional growth."],
    ["2026-02-04", "UBER Q4 2025 Earnings", "FCF (#7)", "N/A", "On Track", "N/A", "5", "FCF $9.8B for FY2025 (+41% YoY). Industry-leading 18.8% FCF margin."],
    ["2026-02-04", "UBER Q4 2025 Earnings", "AV Platform (#2)", "N/A", "On Track", "N/A", "4", "Waymo partnership expanding. 15 AV city target by end-2026 confirmed. $300M Lucid/Nuro investment."],
]

for i, entry in enumerate(log_entries):
    for j, val in enumerate(entry):
        ws2.cell(row=4+i, column=1+j, value=val).font = blk
        ws2.cell(row=4+i, column=1+j).border = bdr

for r in range(4+len(log_entries), 26):
    for c in range(1, 9):
        ws2.cell(row=r, column=c).border = bdr

for i, w in enumerate(log_widths):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# ========== SHEET 3: Rating ==========
ws3 = wb.create_sheet("Rating")
ws3.sheet_properties.tabColor = "3D7AB5"

ws3["A1"] = "Rating & Target Price"
ws3["A1"].font = tf

rating_data = [
    ("Current Rating", "BUY"),
    ("Target Price", "$100"),
    ("Current Price", "$72.83"),
    ("Upside / (Downside)", "+37.3%"),
    ("", ""),
    ("Conviction Score", "=Thesis!H14"),
    ("Pillars On Track", "=Thesis!H15"),
    ("Pillars At Risk", "=Thesis!H16"),
    ("Pillars Broken", "=Thesis!H17"),
    ("", ""),
    ("Last Updated", DATE),
    ("Next Review", "Weekly / post-material event"),
    ("", ""),
    ("Rating Change Triggers:", ""),
    ("Downgrade to HOLD if:", "2+ pillars At Risk AND conviction < 3.0"),
    ("Downgrade to SELL if:", "2+ pillars Broken OR conviction < 2.0"),
    ("Upgrade criteria:", "All pillars On Track AND conviction >= 4.5"),
]

for i, (label, val) in enumerate(rating_data):
    r = 3 + i
    ws3.cell(row=r, column=1, value=label).font = bld if label and ":" not in label else (sf if ":" in label else blk)
    cell = ws3.cell(row=r, column=2, value=val)
    cell.font = Font(name="Arial", bold=True, size=12, color="1B6B3A") if val == "BUY" else blk
    if label == "Current Rating":
        cell.fill = PatternFill("solid", fgColor="C8E6C9")
    cell.border = bdr
    ws3.cell(row=r, column=1).border = bdr

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 45

out = f"coverage/{TICKER}/06-thesis-tracker.xlsx"
wb.save(out)
print(f"Thesis tracker saved to: {out}")
