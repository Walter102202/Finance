import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
wb = openpyxl.Workbook()

hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
black = Font(name="Arial", color="000000", size=10)
green = Font(name="Arial", color="008000", size=10)
bold = Font(name="Arial", bold=True, size=10)
title_font = Font(name="Arial", bold=True, size=14, color="1B3A5C")
sub_font = Font(name="Arial", bold=True, size=11, color="2C5F8A")
alt = PatternFill("solid", fgColor="F2F6FA")
yellow = PatternFill("solid", fgColor="FFFF00")
bdr = Border(left=Side("thin", "CCCCCC"), right=Side("thin", "CCCCCC"), top=Side("thin", "CCCCCC"), bottom=Side("thin", "CCCCCC"))
ctr = Alignment(horizontal="center", vertical="center")

def style_hdr(ws, r, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, ctr, bdr

# ========== SHEET 1: Assumptions ==========
ws = wb.active
ws.title = "Assumptions"
ws.sheet_properties.tabColor = "1B3A5C"

ws["A1"] = f"DCF Model — {TICKER} (Uber Technologies)"
ws["A1"].font = title_font

ws["A3"] = "Key Assumptions"
ws["A3"].font = sub_font

assumptions = [
    ("Current Stock Price ($)", 72.83, "$#,##0.00"),
    ("Shares Outstanding (M)", 2084, "#,##0.0"),
    ("", "", ""),
    ("Risk-Free Rate (10Y UST)", 0.043, "0.00%"),
    ("Equity Risk Premium", 0.055, "0.00%"),
    ("Country Risk Premium", 0.005, "0.00%"),
    ("Levered Beta", 1.15, "0.00"),
    ("Cost of Equity", None, "0.00%"),  # formula
    ("", "", ""),
    ("Pre-Tax Cost of Debt", 0.055, "0.00%"),
    ("Tax Rate", 0.21, "0.00%"),
    ("After-Tax Cost of Debt", None, "0.00%"),  # formula
    ("", "", ""),
    ("Market Cap ($M)", None, "$#,##0"),  # formula
    ("Total Debt ($M)", 9800, "$#,##0"),
    ("Cash ($M)", 7000, "$#,##0"),
    ("Enterprise Value ($M)", None, "$#,##0"),  # formula
    ("Debt Weight", None, "0.0%"),  # formula
    ("Equity Weight", None, "0.0%"),  # formula
    ("WACC", None, "0.00%"),  # formula
    ("", "", ""),
    ("Terminal Growth Rate (Base)", 0.030, "0.0%"),
    ("Exit EV/EBITDA Multiple (Base)", 20, "0.0x"),
]

for i, (label, val, fmt) in enumerate(assumptions):
    r = 4 + i
    ws.cell(row=r, column=1, value=label).font = bold if label else black
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
    if val is not None:
        ws.cell(row=r, column=2, value=val).font = blue
        ws.cell(row=r, column=2).number_format = fmt
        ws.cell(row=r, column=2).fill = yellow
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

# Formulas
ws.cell(row=11, column=2).value = "=B7+B10*(B8+B9)"  # Cost of Equity
ws.cell(row=11, column=2).font = black
ws.cell(row=11, column=2).fill = PatternFill()
ws.cell(row=15, column=2).value = "=B13*(1-B14)"  # After-tax CoD
ws.cell(row=15, column=2).font = black
ws.cell(row=15, column=2).fill = PatternFill()
ws.cell(row=17, column=2).value = "=B4*B5"  # Market cap
ws.cell(row=17, column=2).font = black
ws.cell(row=17, column=2).fill = PatternFill()
ws.cell(row=20, column=2).value = "=B17+B18-B19"  # EV
ws.cell(row=20, column=2).font = black
ws.cell(row=20, column=2).fill = PatternFill()
ws.cell(row=21, column=2).value = "=B18/(B17+B18)"  # Debt weight
ws.cell(row=21, column=2).font = black
ws.cell(row=21, column=2).fill = PatternFill()
ws.cell(row=22, column=2).value = "=B17/(B17+B18)"  # Equity weight
ws.cell(row=22, column=2).font = black
ws.cell(row=22, column=2).fill = PatternFill()
ws.cell(row=23, column=2).value = "=B22*B11+B21*B15"  # WACC
ws.cell(row=23, column=2).font = black
ws.cell(row=23, column=2).fill = PatternFill()

# Revenue growth assumptions
ws["A28"] = "Revenue Growth Assumptions"
ws["A28"].font = sub_font
grow_hdrs = ["", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
for i, h in enumerate(grow_hdrs):
    ws.cell(row=29, column=1+i, value=h)
style_hdr(ws, 29, 7)

ws.cell(row=30, column=1, value="Revenue Growth Rate").font = bold
growth_rates = [0.183, 0.17, 0.15, 0.13, 0.11, 0.10]
for i, g in enumerate(growth_rates):
    ws.cell(row=30, column=2+i, value=g).font = blue
    ws.cell(row=30, column=2+i).number_format = "0.0%"
    ws.cell(row=30, column=2+i).fill = yellow

ws.cell(row=31, column=1, value="EBITDA Margin").font = bold
margins = [0.121, 0.155, 0.175, 0.195, 0.210, 0.220]
for i, m in enumerate(margins):
    ws.cell(row=31, column=2+i, value=m).font = blue
    ws.cell(row=31, column=2+i).number_format = "0.0%"
    ws.cell(row=31, column=2+i).fill = yellow

ws.cell(row=32, column=1, value="Capex % of Revenue").font = bold
capex = [0.035, 0.035, 0.030, 0.030, 0.025, 0.025]
for i, c in enumerate(capex):
    ws.cell(row=32, column=2+i, value=c).font = blue
    ws.cell(row=32, column=2+i).number_format = "0.0%"
    ws.cell(row=32, column=2+i).fill = yellow

ws.cell(row=33, column=1, value="D&A % of Revenue").font = bold
da = [0.025, 0.025, 0.025, 0.025, 0.025, 0.025]
for i, d in enumerate(da):
    ws.cell(row=33, column=2+i, value=d).font = blue
    ws.cell(row=33, column=2+i).number_format = "0.0%"

ws.cell(row=34, column=1, value="Change in WC % of Rev").font = bold
wc = [0.01, 0.01, 0.01, 0.008, 0.008, 0.005]
for i, w in enumerate(wc):
    ws.cell(row=34, column=2+i, value=w).font = blue
    ws.cell(row=34, column=2+i).number_format = "0.0%"

# ========== SHEET 2: FCF Projections ==========
ws2 = wb.create_sheet("FCF Projections")
ws2.sheet_properties.tabColor = "2C5F8A"

ws2["A1"] = f"Free Cash Flow Projections — {TICKER}"
ws2["A1"].font = title_font

fcf_hdrs = ["", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
for i, h in enumerate(fcf_hdrs):
    ws2.cell(row=3, column=1+i, value=h)
style_hdr(ws2, 3, 7)

labels = ["Revenue ($M)", "Growth %", "", "EBITDA ($M)", "EBITDA Margin",
          "(-) D&A ($M)", "EBIT ($M)", "(-) Taxes on EBIT ($M)", "NOPAT ($M)", "",
          "(-) Capex ($M)", "(-) Change in WC ($M)", "(+) D&A ($M)", "Unlevered FCF ($M)"]

for i, l in enumerate(labels):
    ws2.cell(row=4+i, column=1, value=l).font = bold if l and "(" not in l else black
ws2.column_dimensions["A"].width = 25

aref = "Assumptions!"
# 2025A actuals
ws2.cell(row=4, column=2, value=52020).font = blue  # Revenue
ws2.cell(row=4, column=2).number_format = '#,##0'
ws2.cell(row=5, column=2).value = f"={aref}B30"  # Growth
ws2.cell(row=5, column=2).number_format = '0.0%'
ws2.cell(row=7, column=2, value=6310).font = blue  # EBITDA
ws2.cell(row=7, column=2).number_format = '#,##0'
ws2.cell(row=8, column=2).value = "=B7/B4"  # Margin
ws2.cell(row=8, column=2).number_format = '0.0%'

# Projections 2026E-2030E (cols C-G)
for ci in range(5):
    col = 3 + ci
    prev = get_column_letter(col - 1)
    cl = get_column_letter(col)
    acol = get_column_letter(3 + ci)

    ws2.cell(row=4, column=col).value = f"={prev}4*(1+{aref}{acol}30)"
    ws2.cell(row=4, column=col).font = black
    ws2.cell(row=4, column=col).number_format = '#,##0'
    ws2.cell(row=5, column=col).value = f"={aref}{acol}30"
    ws2.cell(row=5, column=col).number_format = '0.0%'
    ws2.cell(row=7, column=col).value = f"={cl}4*{aref}{acol}31"
    ws2.cell(row=7, column=col).font = black
    ws2.cell(row=7, column=col).number_format = '#,##0'
    ws2.cell(row=8, column=col).value = f"={cl}7/{cl}4"
    ws2.cell(row=8, column=col).number_format = '0.0%'
    ws2.cell(row=9, column=col).value = f"={cl}4*{aref}{acol}33"
    ws2.cell(row=9, column=col).font = black
    ws2.cell(row=9, column=col).number_format = '#,##0'
    ws2.cell(row=10, column=col).value = f"={cl}7-{cl}9"
    ws2.cell(row=10, column=col).font = black
    ws2.cell(row=10, column=col).number_format = '#,##0'
    ws2.cell(row=11, column=col).value = f"={cl}10*{aref}B14"
    ws2.cell(row=11, column=col).font = black
    ws2.cell(row=11, column=col).number_format = '#,##0'
    ws2.cell(row=12, column=col).value = f"={cl}10-{cl}11"
    ws2.cell(row=12, column=col).font = black
    ws2.cell(row=12, column=col).number_format = '#,##0'
    ws2.cell(row=14, column=col).value = f"={cl}4*{aref}{acol}32"
    ws2.cell(row=14, column=col).font = black
    ws2.cell(row=14, column=col).number_format = '#,##0'
    ws2.cell(row=15, column=col).value = f"={cl}4*{aref}{acol}34"
    ws2.cell(row=15, column=col).font = black
    ws2.cell(row=15, column=col).number_format = '#,##0'
    ws2.cell(row=16, column=col).value = f"={cl}9"
    ws2.cell(row=16, column=col).font = black
    ws2.cell(row=16, column=col).number_format = '#,##0'
    ws2.cell(row=17, column=col).value = f"={cl}12-{cl}14-{cl}15+{cl}16"
    ws2.cell(row=17, column=col).font = black
    ws2.cell(row=17, column=col).number_format = '#,##0'

for c in range(1, 8):
    ws2.cell(row=17, column=c).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=17, column=c).border = Border(top=Side("medium", "1B3A5C"), bottom=Side("double", "1B3A5C"))

for c in range(2, 8):
    ws2.column_dimensions[get_column_letter(c)].width = 14

# ========== SHEET 3: Valuation ==========
ws3 = wb.create_sheet("Valuation")
ws3.sheet_properties.tabColor = "3D7AB5"

ws3["A1"] = f"DCF Valuation — {TICKER}"
ws3["A1"].font = title_font

ws3["A3"] = "Present Value of Free Cash Flows"
ws3["A3"].font = sub_font

pv_hdrs = ["", "2026E", "2027E", "2028E", "2029E", "2030E"]
for i, h in enumerate(pv_hdrs):
    ws3.cell(row=4, column=1+i, value=h)
style_hdr(ws3, 4, 6)

ws3.cell(row=5, column=1, value="UFCF ($M)").font = bold
ws3.cell(row=6, column=1, value="Discount Factor").font = black
ws3.cell(row=7, column=1, value="PV of FCF ($M)").font = bold

fcf_ref = "'FCF Projections'!"
for i in range(5):
    col = 2 + i
    yr = i + 1
    cl = get_column_letter(col)
    fcf_col = get_column_letter(3 + i)
    ws3.cell(row=5, column=col).value = f"={fcf_ref}{fcf_col}17"
    ws3.cell(row=5, column=col).font = green
    ws3.cell(row=5, column=col).number_format = '#,##0'
    ws3.cell(row=6, column=col).value = f"=1/(1+{aref}B23)^{yr}"
    ws3.cell(row=6, column=col).number_format = '0.000'
    ws3.cell(row=7, column=col).value = f"={cl}5*{cl}6"
    ws3.cell(row=7, column=col).font = black
    ws3.cell(row=7, column=col).number_format = '#,##0'

ws3.cell(row=9, column=1, value="Sum of PV of FCFs ($M)").font = bold
ws3.cell(row=9, column=2).value = "=SUM(B7:F7)"
ws3.cell(row=9, column=2).font = black
ws3.cell(row=9, column=2).number_format = '#,##0'

ws3["A11"] = "Terminal Value"
ws3["A11"].font = sub_font

ws3.cell(row=12, column=1, value="Method 1: Perpetuity Growth").font = bold
ws3.cell(row=13, column=1, value="Terminal FCF (2030E)").font = black
ws3.cell(row=13, column=2).value = f"={fcf_ref}G17"
ws3.cell(row=13, column=2).font = green
ws3.cell(row=13, column=2).number_format = '#,##0'
ws3.cell(row=14, column=1, value="Terminal Growth Rate").font = black
ws3.cell(row=14, column=2).value = f"={aref}B25"
ws3.cell(row=14, column=2).number_format = '0.0%'
ws3.cell(row=15, column=1, value="Terminal Value (Perpetuity)").font = bold
ws3.cell(row=15, column=2).value = f"=B13*(1+B14)/({aref}B23-B14)"
ws3.cell(row=15, column=2).number_format = '#,##0'
ws3.cell(row=16, column=1, value="PV of Terminal Value").font = bold
ws3.cell(row=16, column=2).value = f"=B15/(1+{aref}B23)^5"
ws3.cell(row=16, column=2).number_format = '#,##0'

ws3.cell(row=18, column=1, value="Method 2: Exit Multiple").font = bold
ws3.cell(row=19, column=1, value="Terminal EBITDA (2030E)").font = black
ws3.cell(row=19, column=2).value = f"={fcf_ref}G7"
ws3.cell(row=19, column=2).font = green
ws3.cell(row=19, column=2).number_format = '#,##0'
ws3.cell(row=20, column=1, value="Exit EV/EBITDA Multiple").font = black
ws3.cell(row=20, column=2).value = f"={aref}B26"
ws3.cell(row=20, column=2).number_format = '0.0x'
ws3.cell(row=21, column=1, value="Terminal Value (Exit Multiple)").font = bold
ws3.cell(row=21, column=2).value = "=B19*B20"
ws3.cell(row=21, column=2).number_format = '#,##0'
ws3.cell(row=22, column=1, value="PV of Terminal Value").font = bold
ws3.cell(row=22, column=2).value = f"=B21/(1+{aref}B23)^5"
ws3.cell(row=22, column=2).number_format = '#,##0'

ws3["A24"] = "Implied Valuation"
ws3["A24"].font = sub_font

for i, h in enumerate(["", "Perpetuity Method", "Exit Multiple Method", "Blended (50/50)"]):
    ws3.cell(row=25, column=1+i, value=h)
style_hdr(ws3, 25, 4)

rows_val = [
    ("Sum of PV of FCFs", "=B9", "=B9", "=B9"),
    ("PV of Terminal Value", "=B16", "=B22", "=(B16+B22)/2"),
    ("Enterprise Value ($M)", "=B26+B27", "=C26+C27", "=D26+D27"),
    ("(-) Net Debt ($M)", f"={aref}B18-{aref}B19", f"={aref}B18-{aref}B19", f"={aref}B18-{aref}B19"),
    ("Equity Value ($M)", "=B28-B29", "=C28-C29", "=D28-D29"),
    ("Shares Outstanding (M)", f"={aref}B5", f"={aref}B5", f"={aref}B5"),
    ("Implied Price per Share ($)", "=B30/B31", "=C30/C31", "=D30/D31"),
    ("Current Price ($)", f"={aref}B4", f"={aref}B4", f"={aref}B4"),
    ("Upside / (Downside)", "=(B32-B33)/B33", "=(C32-C33)/C33", "=(D32-D33)/D33"),
]

for i, (label, f1, f2, f3) in enumerate(rows_val):
    r = 26 + i
    ws3.cell(row=r, column=1, value=label).font = bold if "Value" in label or "Price" in label or "Upside" in label else black
    ws3.cell(row=r, column=2).value = f1
    ws3.cell(row=r, column=3).value = f2
    ws3.cell(row=r, column=4).value = f3
    for c in [2, 3, 4]:
        ws3.cell(row=r, column=c).font = black
        ws3.cell(row=r, column=c).border = bdr
        if "Price" in label or "Current" in label:
            ws3.cell(row=r, column=c).number_format = '$#,##0'
        elif "Upside" in label:
            ws3.cell(row=r, column=c).number_format = '0.0%'
        else:
            ws3.cell(row=r, column=c).number_format = '#,##0'

for c in [2, 3, 4]:
    ws3.cell(row=32, column=c).fill = PatternFill("solid", fgColor="E8F5E9")
    ws3.cell(row=32, column=c).font = Font(name="Arial", bold=True, size=11, color="1B3A5C")
    ws3.cell(row=34, column=c).fill = PatternFill("solid", fgColor="FFF3CD")
    ws3.cell(row=34, column=c).font = Font(name="Arial", bold=True, size=10)

ws3.column_dimensions["A"].width = 30
for c in ["B", "C", "D", "E", "F"]:
    ws3.column_dimensions[c].width = 20

# ========== SHEET 4: Sensitivity ==========
ws4 = wb.create_sheet("Sensitivity")
ws4.sheet_properties.tabColor = "4A90D9"

ws4["A1"] = "Sensitivity Analysis"
ws4["A1"].font = title_font

ws4["A3"] = "Implied Price: WACC vs Terminal Growth Rate (Perpetuity)"
ws4["A3"].font = sub_font

wacc_vals = [0.07, 0.08, 0.09, 0.10, 0.11, 0.12]
tg_vals = [0.020, 0.025, 0.030, 0.035, 0.040]

ws4.cell(row=4, column=1, value="WACC \\ TGR").font = hdr_font
ws4.cell(row=4, column=1).fill = hdr_fill
ws4.cell(row=4, column=1).border = bdr
for j, tg in enumerate(tg_vals):
    ws4.cell(row=4, column=2+j, value=tg).font = hdr_font
    ws4.cell(row=4, column=2+j).fill = hdr_fill
    ws4.cell(row=4, column=2+j).number_format = '0.0%'
    ws4.cell(row=4, column=2+j).border = bdr
    ws4.cell(row=4, column=2+j).alignment = ctr

fcf_cells = ["'FCF Projections'!C17", "'FCF Projections'!D17", "'FCF Projections'!E17", "'FCF Projections'!F17", "'FCF Projections'!G17"]
net_debt = f"({aref}B18-{aref}B19)"
shares = f"{aref}B5"

for i, wacc in enumerate(wacc_vals):
    r = 5 + i
    ws4.cell(row=r, column=1, value=wacc).font = bold
    ws4.cell(row=r, column=1).number_format = '0.0%'
    ws4.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D5E8F0")
    ws4.cell(row=r, column=1).border = bdr
    for j, tg in enumerate(tg_vals):
        col = 2 + j
        pv_sum = "+".join([f"{fc}/(1+$A{r})^{k+1}" for k, fc in enumerate(fcf_cells)])
        tv = f"('FCF Projections'!G17*(1+{get_column_letter(col)}$4))/($A{r}-{get_column_letter(col)}$4)/(1+$A{r})^5"
        ws4.cell(row=r, column=col).value = f"=({pv_sum}+{tv}-{net_debt})/{shares}"
        ws4.cell(row=r, column=col).font = black
        ws4.cell(row=r, column=col).number_format = '$#,##0'
        ws4.cell(row=r, column=col).border = bdr
        ws4.cell(row=r, column=col).alignment = ctr

ws4["A13"] = "Implied Price: WACC vs Exit EV/EBITDA Multiple"
ws4["A13"].font = sub_font

exit_mults = [16, 18, 20, 22, 24]
ws4.cell(row=14, column=1, value="WACC \\ Exit Mult").font = hdr_font
ws4.cell(row=14, column=1).fill = hdr_fill
ws4.cell(row=14, column=1).border = bdr
for j, em in enumerate(exit_mults):
    ws4.cell(row=14, column=2+j, value=em).font = hdr_font
    ws4.cell(row=14, column=2+j).fill = hdr_fill
    ws4.cell(row=14, column=2+j).number_format = '0.0x'
    ws4.cell(row=14, column=2+j).border = bdr
    ws4.cell(row=14, column=2+j).alignment = ctr

for i, wacc in enumerate(wacc_vals):
    r = 15 + i
    ws4.cell(row=r, column=1, value=wacc).font = bold
    ws4.cell(row=r, column=1).number_format = '0.0%'
    ws4.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D5E8F0")
    ws4.cell(row=r, column=1).border = bdr
    for j in range(5):
        col = 2 + j
        pv_sum = "+".join([f"{fc}/(1+$A{r})^{k+1}" for k, fc in enumerate(fcf_cells)])
        tv_exit = f"('FCF Projections'!G7*{get_column_letter(col)}$14)/(1+$A{r})^5"
        ws4.cell(row=r, column=col).value = f"=({pv_sum}+{tv_exit}-{net_debt})/{shares}"
        ws4.cell(row=r, column=col).font = black
        ws4.cell(row=r, column=col).number_format = '$#,##0'
        ws4.cell(row=r, column=col).border = bdr
        ws4.cell(row=r, column=col).alignment = ctr

ws4.column_dimensions["A"].width = 18
for c in ["B","C","D","E","F"]:
    ws4.column_dimensions[c].width = 14

# ========== SHEET 5: Scenarios ==========
ws5 = wb.create_sheet("Scenarios")
ws5.sheet_properties.tabColor = "6BB86B"

ws5["A1"] = "Scenario Analysis"
ws5["A1"].font = title_font

scen_hdrs = ["Assumption", "Bear Case", "Base Case", "Bull Case"]
for i, h in enumerate(scen_hdrs):
    ws5.cell(row=3, column=1+i, value=h)
style_hdr(ws5, 3, 4)

scenarios = [
    ("Revenue CAGR (5Y)", "10%", "13%", "16%"),
    ("Terminal EBITDA Margin", "18%", "22%", "26%"),
    ("WACC", "11.0%", "9.2%", "8.0%"),
    ("Terminal Growth Rate", "2.0%", "3.0%", "4.0%"),
    ("Exit EV/EBITDA", "16.0x", "20.0x", "24.0x"),
    ("", "", "", ""),
    ("Implied Price/Share", "", "", ""),
    ("vs Current ($72.83)", "", "", ""),
]

bear_fill = PatternFill("solid", fgColor="FFCDD2")
base_fill = PatternFill("solid", fgColor="FFF3CD")
bull_fill = PatternFill("solid", fgColor="C8E6C9")

for i, (label, bear, base, bull_v) in enumerate(scenarios):
    r = 4 + i
    ws5.cell(row=r, column=1, value=label).font = bold if label else black
    ws5.cell(row=r, column=2, value=bear).font = black
    ws5.cell(row=r, column=2).fill = bear_fill if bear else PatternFill()
    ws5.cell(row=r, column=3, value=base).font = black
    ws5.cell(row=r, column=3).fill = base_fill if base else PatternFill()
    ws5.cell(row=r, column=4, value=bull_v).font = black
    ws5.cell(row=r, column=4).fill = bull_fill if bull_v else PatternFill()
    for c in range(1, 5):
        ws5.cell(row=r, column=c).border = bdr

ws5.cell(row=10, column=1, value="Implied Price/Share").font = Font(name="Arial", bold=True, size=11, color="1B3A5C")
ws5.cell(row=10, column=2, value=50).font = Font(name="Arial", bold=True, size=11, color="FF0000")
ws5.cell(row=10, column=2).number_format = '$#,##0'
ws5.cell(row=10, column=2).fill = bear_fill
ws5.cell(row=10, column=3).value = "=Valuation!D32"
ws5.cell(row=10, column=3).font = Font(name="Arial", bold=True, size=11)
ws5.cell(row=10, column=3).number_format = '$#,##0'
ws5.cell(row=10, column=3).fill = base_fill
ws5.cell(row=10, column=4, value=140).font = Font(name="Arial", bold=True, size=11, color="008000")
ws5.cell(row=10, column=4).number_format = '$#,##0'
ws5.cell(row=10, column=4).fill = bull_fill

ws5.cell(row=11, column=1, value="Upside / (Downside)").font = bold
for c, ref in [(2, "B10"), (3, "C10"), (4, "D10")]:
    ws5.cell(row=11, column=c).value = f"=({ref}-{aref}B4)/{aref}B4"
    ws5.cell(row=11, column=c).number_format = '0.0%'
    ws5.cell(row=11, column=c).border = bdr

ws5.column_dimensions["A"].width = 25
for c in ["B", "C", "D"]:
    ws5.column_dimensions[c].width = 18

out = f"coverage/{TICKER}/03-valuation/dcf-model.xlsx"
wb.save(out)
print(f"DCF model saved to: {out}")
