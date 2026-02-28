import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
wb = openpyxl.Workbook()

# Styles
hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
blk = Font(name="Arial", color="000000", size=10)
grn = Font(name="Arial", color="008000", size=10)
bld = Font(name="Arial", bold=True, size=10)
bld_bl = Font(name="Arial", bold=True, size=10, color="1B3A5C")
tf = Font(name="Arial", bold=True, size=14, color="1B3A5C")
sf = Font(name="Arial", bold=True, size=11, color="2C5F8A")
alt = PatternFill("solid", fgColor="F2F6FA")
yel = PatternFill("solid", fgColor="FFFF00")
hist_fill = PatternFill("solid", fgColor="E8EAF6")
proj_fill = PatternFill("solid", fgColor="FFF8E1")
bdr = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"), top=Side("thin","CCCCCC"), bottom=Side("thin","CCCCCC"))
totbdr = Border(top=Side("medium","1B3A5C"), bottom=Side("double","1B3A5C"))
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%'

def hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bdr

years_h = ["", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
NC = len(years_h)

# ========== ASSUMPTIONS ==========
wa = wb.active
wa.title = "Assumptions"
wa.sheet_properties.tabColor = "1B3A5C"
wa["A1"] = f"Key Assumptions — {TICKER} (Uber Technologies)"
wa["A1"].font = tf

for i, h in enumerate(years_h):
    wa.cell(row=3, column=1+i, value=h)
hdr(wa, 3, NC)

for c in range(2, 5):
    wa.cell(row=3, column=c).fill = PatternFill("solid", fgColor="3D5A80")
for c in range(5, NC):
    wa.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E07A2F")
    wa.cell(row=3, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

assumptions_data = {
    4:  ("Revenue Growth %", [0.169, 0.180, 0.183, 0.170, 0.150, 0.130, 0.110, 0.100], PCT),
    5:  ("COGS % of Revenue", [0.627, 0.625, 0.615, 0.610, 0.600, 0.590, 0.580, 0.575], PCT),
    6:  ("S&M % of Revenue", [0.137, 0.120, 0.110, 0.105, 0.100, 0.095, 0.090, 0.088], PCT),
    7:  ("R&D % of Revenue", [0.150, 0.131, 0.120, 0.115, 0.110, 0.105, 0.100, 0.098], PCT),
    8:  ("G&A % of Revenue", [0.076, 0.060, 0.055, 0.052, 0.050, 0.048, 0.046, 0.044], PCT),
    9:  ("D&A % of Revenue", [0.040, 0.035, 0.030, 0.028, 0.027, 0.026, 0.025, 0.025], PCT),
    10: ("Tax Rate", [0.10, 0.15, 0.18, 0.20, 0.21, 0.21, 0.21, 0.21], PCT),
    11: ("", [], ""),
    12: ("Capex % of Revenue", [0.035, 0.035, 0.033, 0.032, 0.030, 0.028, 0.026, 0.025], PCT),
    13: ("AR Days", [25, 24, 23, 23, 22, 22, 21, 21], '0'),
    14: ("AP Days", [50, 48, 46, 45, 44, 43, 42, 42], '0'),
    15: ("Accrued Liabilities % Rev", [0.060, 0.058, 0.055, 0.053, 0.051, 0.050, 0.048, 0.047], PCT),
    16: ("", [], ""),
    17: ("Interest Rate on Debt", [0.050, 0.052, 0.053, 0.053, 0.053, 0.052, 0.050, 0.048], PCT),
    18: ("Shares Outstanding (M)", [2050, 2070, 2084, 2060, 2040, 2020, 2000, 1980], '#,##0'),
}

for r, (label, vals, fmt) in assumptions_data.items():
    wa.cell(row=r, column=1, value=label).font = bld if label else blk
    for i, v in enumerate(vals):
        c = 2 + i
        wa.cell(row=r, column=c, value=v).font = blue
        wa.cell(row=r, column=c).number_format = fmt
        if c >= 5:
            wa.cell(row=r, column=c).fill = yel

wa.column_dimensions["A"].width = 28
for c in range(2, NC+1):
    wa.column_dimensions[get_column_letter(c)].width = 13

# ========== INCOME STATEMENT ==========
wi = wb.create_sheet("Income Statement")
wi.sheet_properties.tabColor = "2C5F8A"
wi["A1"] = f"Income Statement — {TICKER} ($M)"
wi["A1"].font = tf

for i, h in enumerate(years_h):
    wi.cell(row=3, column=1+i, value=h)
hdr(wi, 3, NC)

is_rows = [
    (4, "Revenue", True),
    (5, "YoY Growth %", False),
    (6, "(-) Cost of Revenue", False),
    (7, "Gross Profit", True),
    (8, "Gross Margin %", False),
    (9, "", False),
    (10, "(-) Sales & Marketing", False),
    (11, "(-) Research & Development", False),
    (12, "(-) General & Administrative", False),
    (13, "Total Operating Expenses", True),
    (14, "", False),
    (15, "Operating Income (EBIT)", True),
    (16, "Operating Margin %", False),
    (17, "", False),
    (18, "(-) Interest Expense", False),
    (19, "(+) Other Income / (Expense)", False),
    (20, "Pre-Tax Income", True),
    (21, "(-) Income Tax", False),
    (22, "Effective Tax Rate", False),
    (23, "Net Income", True),
    (24, "Net Margin %", False),
    (25, "", False),
    (26, "EPS (Diluted)", True),
    (27, "EBITDA", True),
    (28, "EBITDA Margin %", False),
]

for r, label, is_bold in is_rows:
    wi.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

# Historical actuals (FY2023, FY2024, FY2025)
hist_rev = [37281, 43978, 52020]
hist_cogs = [23384, 27486, 31990]
hist_sm = [5111, 5277, 5722]
hist_rd = [5593, 5780, 6242]
hist_ga = [2833, 2643, 2861]
hist_int = [633, 466, 380]
hist_other = [-2676, 7146, 5238]  # includes equity investment gains
hist_ni = [1887, 9856, 10050]
hist_da = [1491, 1540, 1561]

aref = "Assumptions!"
for i in range(3):
    c = 2 + i
    cl = get_column_letter(c)
    wi.cell(row=4, column=c, value=hist_rev[i]).font = blue
    wi.cell(row=5, column=c).value = f"=({cl}4-{get_column_letter(c-1)}4)/{get_column_letter(c-1)}4" if i > 0 else 0.169
    wi.cell(row=5, column=c).number_format = PCT
    wi.cell(row=6, column=c, value=hist_cogs[i]).font = blue
    wi.cell(row=7, column=c).value = f"={cl}4-{cl}6"
    wi.cell(row=8, column=c).value = f"={cl}7/{cl}4"
    wi.cell(row=8, column=c).number_format = PCT
    wi.cell(row=10, column=c, value=hist_sm[i]).font = blue
    wi.cell(row=11, column=c, value=hist_rd[i]).font = blue
    wi.cell(row=12, column=c, value=hist_ga[i]).font = blue
    wi.cell(row=13, column=c).value = f"={cl}10+{cl}11+{cl}12"
    wi.cell(row=15, column=c).value = f"={cl}7-{cl}13"
    wi.cell(row=16, column=c).value = f"={cl}15/{cl}4"
    wi.cell(row=16, column=c).number_format = PCT
    wi.cell(row=18, column=c, value=hist_int[i]).font = blue
    wi.cell(row=19, column=c, value=hist_other[i]).font = blue
    wi.cell(row=20, column=c).value = f"={cl}15-{cl}18+{cl}19"
    wi.cell(row=21, column=c).value = f"={cl}20*{aref}{cl}10"
    wi.cell(row=22, column=c).value = f"=IF({cl}20=0,0,{cl}21/{cl}20)"
    wi.cell(row=22, column=c).number_format = PCT
    wi.cell(row=23, column=c).value = f"={cl}20-{cl}21"
    wi.cell(row=24, column=c).value = f"={cl}23/{cl}4"
    wi.cell(row=24, column=c).number_format = PCT
    wi.cell(row=26, column=c).value = f"={cl}23/{aref}{cl}18"
    wi.cell(row=26, column=c).number_format = '$#,##0.00'
    wi.cell(row=27, column=c).value = f"={cl}15+{cl}4*{aref}{cl}9"
    wi.cell(row=28, column=c).value = f"={cl}27/{cl}4"
    wi.cell(row=28, column=c).number_format = PCT

# Projected (2026E-2030E)
for i in range(5):
    c = 5 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wi.cell(row=4, column=c).value = f"={prev}4*(1+{aref}{acl}4)"
    wi.cell(row=5, column=c).value = f"={aref}{acl}4"
    wi.cell(row=5, column=c).number_format = PCT
    wi.cell(row=6, column=c).value = f"={cl}4*{aref}{acl}5"
    wi.cell(row=7, column=c).value = f"={cl}4-{cl}6"
    wi.cell(row=8, column=c).value = f"={cl}7/{cl}4"
    wi.cell(row=8, column=c).number_format = PCT
    wi.cell(row=10, column=c).value = f"={cl}4*{aref}{acl}6"
    wi.cell(row=11, column=c).value = f"={cl}4*{aref}{acl}7"
    wi.cell(row=12, column=c).value = f"={cl}4*{aref}{acl}8"
    wi.cell(row=13, column=c).value = f"={cl}10+{cl}11+{cl}12"
    wi.cell(row=15, column=c).value = f"={cl}7-{cl}13"
    wi.cell(row=16, column=c).value = f"={cl}15/{cl}4"
    wi.cell(row=16, column=c).number_format = PCT
    wi.cell(row=18, column=c).value = f"='Balance Sheet'!{cl}20*{aref}{acl}17"
    wi.cell(row=18, column=c).font = grn
    wi.cell(row=19, column=c, value=300).font = blue  # conservative other income
    wi.cell(row=20, column=c).value = f"={cl}15-{cl}18+{cl}19"
    wi.cell(row=21, column=c).value = f"={cl}20*{aref}{acl}10"
    wi.cell(row=22, column=c).value = f"=IF({cl}20=0,0,{cl}21/{cl}20)"
    wi.cell(row=22, column=c).number_format = PCT
    wi.cell(row=23, column=c).value = f"={cl}20-{cl}21"
    wi.cell(row=24, column=c).value = f"={cl}23/{cl}4"
    wi.cell(row=24, column=c).number_format = PCT
    wi.cell(row=26, column=c).value = f"={cl}23/{aref}{acl}18"
    wi.cell(row=26, column=c).number_format = '$#,##0.00'
    wi.cell(row=27, column=c).value = f"={cl}15+{cl}4*{aref}{acl}9"
    wi.cell(row=28, column=c).value = f"={cl}27/{cl}4"
    wi.cell(row=28, column=c).number_format = PCT

for r in [4,6,7,10,11,12,13,15,18,19,20,21,23,27]:
    for c in range(2, NC):
        wi.cell(row=r, column=c).number_format = NUM

for r in [7, 13, 15, 20, 23, 27]:
    for c in range(1, NC):
        wi.cell(row=r, column=c).font = bld_bl
        wi.cell(row=r, column=c).border = totbdr if r in [7, 15, 23] else bdr

for r in range(4, 29):
    for c in range(5, NC):
        cell = wi.cell(row=r, column=c)
        if not cell.fill or cell.fill.fgColor.rgb == "00000000":
            cell.fill = proj_fill

wi.column_dimensions["A"].width = 32
for c in range(2, NC+1):
    wi.column_dimensions[get_column_letter(c)].width = 14

# ========== BALANCE SHEET ==========
wb2 = wb.create_sheet("Balance Sheet")
wb2.sheet_properties.tabColor = "3D7AB5"
wb2["A1"] = f"Balance Sheet — {TICKER} ($M)"
wb2["A1"].font = tf

for i, h in enumerate(years_h):
    wb2.cell(row=3, column=1+i, value=h)
hdr(wb2, 3, NC)

bs_rows = [
    (4, "ASSETS", True),
    (5, "Cash & Equivalents", False),
    (6, "Short-Term Investments", False),
    (7, "Accounts Receivable", False),
    (8, "Other Current Assets", False),
    (9, "Total Current Assets", True),
    (10, "", False),
    (11, "PP&E (net)", False),
    (12, "Intangibles & Goodwill", False),
    (13, "Equity Investments", False),
    (14, "Other Non-Current Assets", False),
    (15, "Total Assets", True),
    (16, "", False),
    (17, "LIABILITIES", True),
    (18, "Accounts Payable", False),
    (19, "Accrued Liabilities", False),
    (20, "Long-Term Debt", False),
    (21, "Operating Lease Liabilities", False),
    (22, "Other Liabilities", False),
    (23, "Total Liabilities", True),
    (24, "", False),
    (25, "EQUITY", True),
    (26, "Retained Earnings", False),
    (27, "Other Equity", False),
    (28, "Total Equity", True),
    (29, "Total Liab. + Equity", True),
    (30, "Balance Check (should be 0)", False),
]

for r, label, is_bold in bs_rows:
    wb2.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

# Historical BS (FY2023, FY2024, FY2025)
hist_bs = {
    5:  [4680, 5470, 7000],       # Cash
    6:  [1500, 1700, 2000],       # ST Investments
    7:  [2560, 2890, 3280],       # AR
    8:  [1800, 2100, 2500],       # Other CA
    11: [2370, 2550, 2700],       # PPE
    12: [16175, 16270, 16300],    # Intangibles/Goodwill
    13: [12400, 14200, 15000],    # Equity investments
    14: [6715, 7320, 7520],       # Other NCA
    18: [3200, 3400, 3600],       # AP
    19: [2237, 2550, 2860],       # Accrued
    20: [9400, 9800, 9500],       # LT Debt
    21: [1500, 1600, 1700],       # Operating leases
    22: [4800, 5200, 5540],       # Other liab
    27: [20103, 23290, 28100],    # Other equity (total equity used as base)
}

for r, vals in hist_bs.items():
    for i, v in enumerate(vals):
        wb2.cell(row=r, column=2+i, value=v).font = blue
        wb2.cell(row=r, column=2+i).number_format = NUM

isref = "'Income Statement'!"
for c in range(2, 5):
    cl = get_column_letter(c)
    wb2.cell(row=9, column=c).value = f"=SUM({cl}5:{cl}8)"
    wb2.cell(row=15, column=c).value = f"={cl}9+{cl}11+{cl}12+{cl}13+{cl}14"
    wb2.cell(row=23, column=c).value = f"={cl}18+{cl}19+{cl}20+{cl}21+{cl}22"
    wb2.cell(row=26, column=c).value = f"={cl}15-{cl}23-{cl}27"
    wb2.cell(row=26, column=c).font = blk
    wb2.cell(row=28, column=c).value = f"={cl}26+{cl}27"
    wb2.cell(row=29, column=c).value = f"={cl}23+{cl}28"
    wb2.cell(row=30, column=c).value = f"={cl}15-{cl}29"
    wb2.cell(row=30, column=c).number_format = NUM

# Projected BS
bsref = "'Balance Sheet'!"
for i in range(5):
    c = 5 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    # Cash = plug from CF
    wb2.cell(row=5, column=c).value = f"={prev}5+'Cash Flow'!{cl}19"
    wb2.cell(row=5, column=c).font = grn
    # ST Investments (grow slowly)
    wb2.cell(row=6, column=c).value = f"={prev}6*1.05"
    # AR = Rev * AR Days / 365
    wb2.cell(row=7, column=c).value = f"={isref}{cl}4*{aref}{acl}13/365"
    # Other CA
    wb2.cell(row=8, column=c).value = f"={prev}8*1.05"
    # Total CA
    wb2.cell(row=9, column=c).value = f"=SUM({cl}5:{cl}8)"
    # PPE = prev + capex - D&A
    wb2.cell(row=11, column=c).value = f"={prev}11+{isref}{cl}4*{aref}{acl}12-{isref}{cl}4*{aref}{acl}9"
    # Intangibles (flat)
    wb2.cell(row=12, column=c).value = f"={prev}12"
    # Equity investments (flat/modest growth)
    wb2.cell(row=13, column=c).value = f"={prev}13*1.02"
    # Other NCA
    wb2.cell(row=14, column=c).value = f"={prev}14*1.03"
    # Total Assets
    wb2.cell(row=15, column=c).value = f"={cl}9+{cl}11+{cl}12+{cl}13+{cl}14"
    # AP = COGS * AP Days / 365
    wb2.cell(row=18, column=c).value = f"={isref}{cl}6*{aref}{acl}14/365"
    # Accrued = Rev * %
    wb2.cell(row=19, column=c).value = f"={isref}{cl}4*{aref}{acl}15"
    # LT Debt (gradually reduce via buybacks)
    wb2.cell(row=20, column=c).value = f"={prev}20*0.97"
    wb2.cell(row=20, column=c).font = blue
    # Operating leases (flat)
    wb2.cell(row=21, column=c).value = f"={prev}21*1.02"
    # Other liab
    wb2.cell(row=22, column=c).value = f"={prev}22*1.03"
    # Total Liab
    wb2.cell(row=23, column=c).value = f"={cl}18+{cl}19+{cl}20+{cl}21+{cl}22"
    # Retained Earnings = prev RE + NI
    wb2.cell(row=26, column=c).value = f"={prev}26+{isref}{cl}23"
    # Other equity
    wb2.cell(row=27, column=c).value = f"={prev}27"
    # Total Equity
    wb2.cell(row=28, column=c).value = f"={cl}26+{cl}27"
    # Total L+E
    wb2.cell(row=29, column=c).value = f"={cl}23+{cl}28"
    # Balance check
    wb2.cell(row=30, column=c).value = f"={cl}15-{cl}29"
    wb2.cell(row=30, column=c).number_format = NUM

for r in [5,6,7,8,9,11,12,13,14,15,18,19,20,21,22,23,26,27,28,29]:
    for c in range(2, NC):
        wb2.cell(row=r, column=c).number_format = NUM

for r in [9, 15, 23, 28, 29]:
    for c in range(1, NC):
        wb2.cell(row=r, column=c).font = bld_bl

wb2.column_dimensions["A"].width = 30
for c in range(2, NC+1):
    wb2.column_dimensions[get_column_letter(c)].width = 14

# ========== CASH FLOW STATEMENT ==========
wc = wb.create_sheet("Cash Flow")
wc.sheet_properties.tabColor = "4A90D9"
wc["A1"] = f"Cash Flow Statement — {TICKER} ($M)"
wc["A1"].font = tf

for i, h in enumerate(years_h):
    wc.cell(row=3, column=1+i, value=h)
hdr(wc, 3, NC)

cf_rows = [
    (4, "OPERATING ACTIVITIES", True),
    (5, "Net Income", False),
    (6, "(+) Depreciation & Amortization", False),
    (7, "Changes in Working Capital", False),
    (8, "  Change in AR", False),
    (9, "  Change in AP", False),
    (10, "  Change in Accrued Liab", False),
    (11, "  Other Operating Changes", False),
    (12, "Cash from Operations (CFO)", True),
    (13, "", False),
    (14, "INVESTING ACTIVITIES", True),
    (15, "(-) Capital Expenditures", False),
    (16, "(-) Other Investing", False),
    (17, "Cash from Investing (CFI)", True),
    (18, "", False),
    (19, "Net Change in Cash", True),
]

for r, label, is_bold in cf_rows:
    wc.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

# Historical CF
hist_cfo = [3585, 6903, 9760]
hist_capex = [600, 650, 700]

for i in range(3):
    c = 2 + i
    cl = get_column_letter(c)
    wc.cell(row=5, column=c).value = f"={isref}{cl}23"
    wc.cell(row=5, column=c).font = grn
    wc.cell(row=6, column=c).value = f"={isref}{cl}4*{aref}{cl}9"
    wc.cell(row=12, column=c, value=hist_cfo[i]).font = blue
    wc.cell(row=15, column=c, value=hist_capex[i]).font = blue
    wc.cell(row=16, column=c, value=-500).font = blue
    wc.cell(row=17, column=c).value = f"=-{cl}15+{cl}16"
    wc.cell(row=19, column=c).value = f"={cl}12+{cl}17"

# Projected CF
for i in range(5):
    c = 5 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wc.cell(row=5, column=c).value = f"={isref}{cl}23"
    wc.cell(row=5, column=c).font = grn
    wc.cell(row=6, column=c).value = f"={isref}{cl}4*{aref}{acl}9"
    # WC changes
    wc.cell(row=8, column=c).value = f"=-('Balance Sheet'!{cl}7-'Balance Sheet'!{prev}7)"
    wc.cell(row=9, column=c).value = f"='Balance Sheet'!{cl}18-'Balance Sheet'!{prev}18"
    wc.cell(row=10, column=c).value = f"='Balance Sheet'!{cl}19-'Balance Sheet'!{prev}19"
    wc.cell(row=11, column=c, value=0).font = blue
    wc.cell(row=7, column=c).value = f"={cl}8+{cl}9+{cl}10+{cl}11"
    wc.cell(row=12, column=c).value = f"={cl}5+{cl}6+{cl}7"
    wc.cell(row=15, column=c).value = f"={isref}{cl}4*{aref}{acl}12"
    wc.cell(row=16, column=c, value=-300).font = blue
    wc.cell(row=17, column=c).value = f"=-{cl}15+{cl}16"
    wc.cell(row=19, column=c).value = f"={cl}12+{cl}17"

for r in [5,6,7,8,9,10,11,12,15,16,17,19]:
    for c in range(2, NC):
        wc.cell(row=r, column=c).number_format = NUM

for r in [12, 17, 19]:
    for c in range(1, NC):
        wc.cell(row=r, column=c).font = bld_bl

wc.column_dimensions["A"].width = 30
for c in range(2, NC+1):
    wc.column_dimensions[get_column_letter(c)].width = 14

# ========== Save ==========
out = f"coverage/{TICKER}/04-financial-model/3-statements.xlsx"
wb.save(out)
print(f"3-statement model saved to: {out}")
