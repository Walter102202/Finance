# DISCLAIMER: Financial data hardcoded in this script is illustrative only (as of Feb 2026). Update with current data before use. Not investment advice.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, datetime

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
QUARTER = sys.argv[2] if len(sys.argv) > 2 else "Q4-2025"
DATE = "2026-02-28"
wb = openpyxl.Workbook()

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
blk = Font(name="Arial", color="000000", size=10)
bld = Font(name="Arial", bold=True, size=10)
bld_bl = Font(name="Arial", bold=True, size=10, color="1B3A5C")
tf = Font(name="Arial", bold=True, size=14, color="1B3A5C")
sf = Font(name="Arial", bold=True, size=11, color="2C5F8A")
bdr = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"), top=Side("thin","CCCCCC"), bottom=Side("thin","CCCCCC"))
totbdr = Border(top=Side("medium","1B3A5C"), bottom=Side("double","1B3A5C"))
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
greenF = PatternFill("solid", fgColor="E8F5E9")
redF = PatternFill("solid", fgColor="FFEBEE")
yellowF = PatternFill("solid", fgColor="FFF8E1")
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%'
DLR = '$#,##0'

def hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bdr

# ========== SHEET 1: TARGET PRICE BRIDGE ==========
ws = wb.active
ws.title = "Target Price Bridge"
ws.sheet_properties.tabColor = "1B3A5C"

ws["A1"] = f"Model Update — {TICKER} | {QUARTER}"
ws["A1"].font = tf
ws["A2"] = f"Date: {DATE} | Post-Earnings Model Revision"
ws["A2"].font = Font(name="Arial", italic=True, color="888888", size=10)

headers = ["Component", "Impact on Target ($)", "Notes"]
widths = [35, 20, 40]
for i, h in enumerate(headers):
    ws.cell(row=4, column=1+i, value=h)
hdr(ws, 4, 3)

bridge_data = [
    ("Previous Target Price", "$100", "Pre-Q4 2025 earnings target (initiation)"),
    ("", "", ""),
    ("Revenue Beat (+$120M vs model)", "+$3", "Q4 rev $14.37B vs model $14.25B; FY rev $52.02B vs model $52.0B"),
    ("Gross Bookings Beat (+$1.0B)", "+$4", "GBs $54.14B vs consensus $53.1B; shows demand strength"),
    ("Adj. EBITDA Beat (+$40M vs model)", "+$2", "Q4 EBITDA $2.49B vs model $2.45B; margin 4.6% of GBs"),
    ("Delivery Margin Expansion (4.0%)", "+$5", "Delivery EBITDA margin to 4.0% from 3.6%; validates long-term thesis"),
    ("AV Platform Momentum (15 cities)", "+$3", "15-city target, Waabi exclusivity, 30% utilization premium"),
    ("Q1 2026 EBITDA Guidance Below Street", "-$6", "Guide $2.37-2.47B vs Street $2.55B; near-term margin pressure"),
    ("2026E EPS Estimate Cut ($4.15->$3.30)", "-$5", "Street-wide EPS cut; lower near-term earnings power"),
    ("Record Capital Returns ($1.9B Q4)", "+$2", "$20B auth; ~50% FCF return target; 2% share count reduction"),
    ("Uber One 46M (+55% YoY)", "+$2", "Approaching 50% of GBs; enhances platform stickiness"),
    ("CFO Transition Risk", "-$1", "Mahendra-Rajah departure; Krishnamurthy internal but new to role"),
    ("GAAP Equity Investment Headwind", "$0", "Non-cash $1.6B revaluation; does not affect operating value"),
    ("", "", ""),
    ("Net Adjustment", "+$9", ""),
    ("", "", ""),
    ("New Target Price", "$109", "Rounded; operational strength offset by near-term margin concern"),
    ("Current Price (Feb 28)", "$74.80", "Post-earnings selloff of ~6.4%"),
    ("Implied Upside", "+45.7%", "Increased from +37% due to price decline + target increase"),
]

for i, (comp, impact, notes) in enumerate(bridge_data):
    r = 5 + i
    ws.cell(row=r, column=1, value=comp).font = bld_bl if comp in ["Previous Target Price", "Net Adjustment", "New Target Price", "Current Price (Feb 28)", "Implied Upside"] else blk
    ws.cell(row=r, column=2, value=impact).font = bld if comp in ["Previous Target Price", "New Target Price", "Implied Upside"] else blk
    ws.cell(row=r, column=3, value=notes).font = blk
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = bdr
        ws.cell(row=r, column=c).alignment = lft if c != 2 else ctr
    if impact.startswith("+$") or impact.startswith("+4"):
        ws.cell(row=r, column=2).fill = greenF
    elif impact.startswith("-$"):
        ws.cell(row=r, column=2).fill = redF
    if comp in ["Previous Target Price", "New Target Price"]:
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = totbdr

for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

# ========== SHEET 2: ASSUMPTIONS COMPARISON ==========
ws2 = wb.create_sheet("Assumptions Update")
ws2.sheet_properties.tabColor = "2C5F8A"

ws2["A1"] = f"Assumptions Update — {TICKER}"
ws2["A1"].font = tf

headers2 = ["Assumption", "2025 Old (E)", "2025 Actual", "Delta", "2026 Old (E)", "2026 New (E)", "Delta", "Notes"]
widths2 = [28, 14, 14, 12, 14, 14, 12, 35]
for i, h in enumerate(headers2):
    ws2.cell(row=3, column=1+i, value=h)
hdr(ws2, 3, len(headers2))

assumptions_update = [
    ("Revenue ($B)", "$52.0B", "$52.02B", "+0.0%", "$60.8B", "$61.3B", "+0.8%", "Higher base; slightly lifted growth"),
    ("Revenue Growth %", "18.3%", "18.3%", "0.0pp", "17.0%", "17.8%", "+0.8pp", "GB momentum supports higher growth"),
    ("Gross Bookings ($B)", "$193B", "$193B", "0.0pp", "$225B", "$228B", "+1.3%", "Q4 beat suggests higher trajectory"),
    ("Adj. EBITDA ($B)", "$8.7B", "$8.7B", "0.0pp", "$10.5B", "$10.2B", "-2.9%", "Lower due to Q1 guide; offset later in year"),
    ("EBITDA Margin % (of GBs)", "4.5%", "4.5%", "0.0pp", "4.7%", "4.5%", "-0.2pp", "Near-term investment compresses margins"),
    ("COGS % of Revenue", "62.0%", "61.8%", "-0.2pp", "61.0%", "61.5%", "+0.5pp", "Higher driver incentives in Q1 2026"),
    ("S&M % of Revenue", "11.0%", "10.8%", "-0.2pp", "10.5%", "11.0%", "+0.5pp", "Investment in affordable products"),
    ("R&D % of Revenue", "12.0%", "11.7%", "-0.3pp", "11.5%", "11.5%", "0.0pp", "Stable; AV investment continues"),
    ("G&A % of Revenue", "5.5%", "5.3%", "-0.2pp", "5.0%", "5.2%", "+0.2pp", "CFO transition; slightly elevated"),
    ("GAAP Op. Margin %", "11.5%", "12.2%", "+0.7pp", "13.0%", "12.5%", "-0.5pp", "GAAP improvement; near-term guide concern"),
    ("Non-GAAP EPS (Annual)", "$2.90", "$2.88", "-0.7%", "$4.15", "$3.30", "-20.5%", "Major EPS estimate reduction by Street"),
    ("FCF ($B)", "$9.5B", "$9.8B", "+3.2%", "$11.0B", "$11.5B", "+4.5%", "FCF stronger; operating leverage"),
    ("MAPCs (M)", "200M", "202M", "+1.0%", "230M", "235M", "+2.2%", "Higher base from Q4 beat"),
    ("Uber One Subs (M)", "43M", "46M", "+7.0%", "55M", "60M", "+9.1%", "55% YoY growth rate suggests higher"),
    ("Ads ARR ($B)", "$1.5B", "$1.5B+", "in line", "$2.0B", "$2.0B", "0.0%", "On track for $2B in 2026"),
    ("Buyback ($B)", "$5.0B", "$5.5B", "+10.0%", "$5.0B", "$5.5B", "+10.0%", "Record pace; ~50% of FCF"),
]

for i, (asn, old25, act25, d25, old26, new26, d26, notes) in enumerate(assumptions_update):
    r = 4 + i
    row_data = [asn, old25, act25, d25, old26, new26, d26, notes]
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=r, column=1+j, value=val)
        cell.font = blk
        cell.alignment = ctr if j > 0 and j < 7 else lft
        cell.border = bdr
    # Highlight deltas
    if "+" in d25 and "pp" not in d25 and d25 != "+0.0%":
        ws2.cell(row=r, column=4).fill = greenF
    elif "-" in d25 and "pp" not in d25:
        ws2.cell(row=r, column=4).fill = redF
    if d25.endswith("pp"):
        if d25.startswith("+"):
            ws2.cell(row=r, column=4).fill = redF if asn in ["COGS % of Revenue", "S&M % of Revenue", "G&A % of Revenue"] else greenF
        elif d25.startswith("-"):
            ws2.cell(row=r, column=4).fill = greenF if asn in ["COGS % of Revenue", "S&M % of Revenue", "G&A % of Revenue"] else redF

for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# ========== SHEET 3: UPDATED KEY METRICS ==========
ws3 = wb.create_sheet("Updated Projections")
ws3.sheet_properties.tabColor = "3D7AB5"

ws3["A1"] = f"Updated Projections — {TICKER} ($M unless noted)"
ws3["A1"].font = tf

years = ["", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E"]
NC = len(years)
for i, h in enumerate(years):
    ws3.cell(row=3, column=1+i, value=h)
hdr(ws3, 3, NC)
for c in range(2, 4):
    ws3.cell(row=3, column=c).fill = PatternFill("solid", fgColor="3D5A80")
for c in range(4, NC):
    ws3.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E07A2F")
    ws3.cell(row=3, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

proj_data = [
    ("Revenue ($M)", [43953, 52020, 61300, 70800, 80400, 89200], NUM),
    ("Revenue Growth %", [0.177, 0.183, 0.178, 0.155, 0.136, 0.109], PCT),
    ("Gross Bookings ($B)", [158, 193, 228, 268, 308, 345], NUM),
    ("GB Growth %", [0.189, 0.222, 0.181, 0.175, 0.149, 0.120], PCT),
    ("", [], ""),
    ("Mobility Revenue ($M)", [25110, 29670, 34420, 39130, 43830, 48210], NUM),
    ("Delivery Revenue ($M)", [13760, 17250, 21330, 25390, 29380, 32910], NUM),
    ("Freight Revenue ($M)", [5083, 5100, 5550, 6280, 7190, 8080], NUM),
    ("", [], ""),
    ("Adj. EBITDA ($M)", [6443, 8700, 10200, 12600, 15200, 17800], NUM),
    ("EBITDA Margin (% Rev)", [0.147, 0.167, 0.166, 0.178, 0.189, 0.200], PCT),
    ("EBITDA Margin (% GBs)", [0.041, 0.045, 0.045, 0.047, 0.049, 0.052], PCT),
    ("", [], ""),
    ("GAAP Operating Income", [2799, 6342, 7663, 9558, 11656, 13784], NUM),
    ("GAAP Op Margin %", [0.064, 0.122, 0.125, 0.135, 0.145, 0.155], PCT),
    ("", [], ""),
    ("Non-GAAP EPS ($)", [2.27, 2.88, 3.30, 4.25, 5.30, 6.40], '$#,##0.00'),
    ("GAAP EPS ($)", [3.45, 3.04, 3.68, 4.58, 5.59, 6.61], '$#,##0.00'),
    ("", [], ""),
    ("FCF ($M)", [6895, 9800, 11500, 13800, 16500, 19200], NUM),
    ("FCF Margin %", [0.157, 0.188, 0.188, 0.195, 0.205, 0.215], PCT),
    ("Buybacks ($M)", [4320, 5500, 5500, 6000, 7000, 8000], NUM),
    ("", [], ""),
    ("MAPCs (M)", [171, 202, 235, 268, 300, 330], NUM),
    ("Trips (B)", [11.2, 13.6, 16.2, 19.0, 21.8, 24.5], '#,##0.0'),
    ("Uber One Subs (M)", [30, 46, 60, 75, 90, 105], NUM),
    ("Ads ARR ($B)", [0.9, 1.5, 2.0, 2.7, 3.4, 4.2], '#,##0.0'),
]

for i, (label, vals, fmt) in enumerate(proj_data):
    r = 4 + i
    ws3.cell(row=r, column=1, value=label).font = bld_bl if label in ["Revenue ($M)", "Adj. EBITDA ($M)", "GAAP Operating Income", "FCF ($M)", "Non-GAAP EPS ($)"] else blk
    for j, v in enumerate(vals):
        cell = ws3.cell(row=r, column=2+j, value=v)
        cell.font = blue if j < 2 else blk
        cell.number_format = fmt
        cell.alignment = ctr
        cell.border = bdr
        if j >= 2:
            cell.fill = yellowF

ws3.column_dimensions["A"].width = 28
for c in range(2, NC+1):
    ws3.column_dimensions[get_column_letter(c)].width = 14

# ========== SHEET 4: DCF REVISION SUMMARY ==========
ws4 = wb.create_sheet("DCF Revision")
ws4.sheet_properties.tabColor = "4A90D9"

ws4["A1"] = "DCF Revision Summary"
ws4["A1"].font = tf

dcf_data = [
    ("WACC", "9.2%", "9.2%", "No change"),
    ("Terminal Growth Rate", "3.0%", "3.0%", "No change"),
    ("Exit Multiple (EV/EBITDA)", "20.0x", "19.5x", "Slight compression due to near-term margin uncertainty"),
    ("", "", "", ""),
    ("PV of FCFs (2026-2030)", "$43,200M", "$44,100M", "+2.1% from higher revenue base and FCF trajectory"),
    ("Terminal Value (Perpetuity)", "$195,000M", "$198,500M", "+1.8% from higher terminal FCF"),
    ("Terminal Value (Exit Multiple)", "$178,000M", "$175,500M", "-1.4% lower multiple offset by higher EBITDA"),
    ("PV of Terminal (Perpetuity)", "$128,700M", "$131,000M", "+1.8%"),
    ("PV of Terminal (Exit)", "$117,500M", "$115,900M", "-1.4%"),
    ("", "", "", ""),
    ("Enterprise Value (Perpetuity)", "$171,900M", "$175,100M", "+1.9%"),
    ("Enterprise Value (Exit)", "$160,700M", "$160,000M", "-0.4%"),
    ("Blended EV (50/50)", "$166,300M", "$167,550M", "+0.8%"),
    ("(-) Net Debt", "$2,500M", "$2,500M", "Held constant ($9.5B debt - $7.0B cash)"),
    ("Equity Value", "$163,800M", "$165,050M", "+0.8%"),
    ("Shares Outstanding (M)", "2,084", "2,042", "2% reduction from buybacks"),
    ("Implied Share Price (Perpetuity)", "$81.30", "$84.50", "+3.9%"),
    ("Implied Share Price (Exit)", "$75.92", "$77.11", "+1.6%"),
    ("Blended DCF Value", "$78.61", "$80.80", "+2.8%"),
    ("", "", "", ""),
    ("Comps-Implied Value", "$120.00", "$125.00", "Higher from delivery margin expansion + AV premium"),
    ("Blended Target (40% Comps / 60% DCF)", "$95.17", "$98.48", "+3.5%"),
    ("Premium for AV Optionality", "+$5", "+$10", "Raised: 15-city target, Waabi exclusivity, 30% utilization"),
    ("Rounded Target Price", "$100", "$109", "+$9 (+9.0%)"),
]

headers4 = ["Item", "Previous", "Updated", "Notes"]
widths4 = [32, 16, 16, 45]
for i, h in enumerate(headers4):
    ws4.cell(row=3, column=1+i, value=h)
hdr(ws4, 3, 4)

for i, (item, prev, upd, notes) in enumerate(dcf_data):
    r = 4 + i
    ws4.cell(row=r, column=1, value=item).font = bld_bl if item in ["Blended EV (50/50)", "Equity Value", "Blended DCF Value", "Rounded Target Price", "Blended Target (40% Comps / 60% DCF)"] else blk
    ws4.cell(row=r, column=2, value=prev).font = blk
    ws4.cell(row=r, column=3, value=upd).font = bld if item == "Rounded Target Price" else blk
    ws4.cell(row=r, column=4, value=notes).font = blk
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = bdr
        ws4.cell(row=r, column=c).alignment = ctr if c in [2, 3] else lft

for i, w in enumerate(widths4):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

out = f"coverage/{TICKER}/08-earnings/{QUARTER}/model-update.xlsx"
wb.save(out)
print(f"Model update saved to: {out}")

# ========== REGENERATE 3-STATEMENTS WITH UPDATED DATA ==========
import subprocess, os
script_dir = os.path.dirname(os.path.abspath(__file__))
update_script = os.path.join(script_dir, "generate-3statements-updated.py")

updated_code = '''# DISCLAIMER: Financial data hardcoded in this script is illustrative only (as of Feb 2026). Update with current data before use. Not investment advice.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
wb = openpyxl.Workbook()

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
blk = Font(name="Arial", color="000000", size=10)
grn = Font(name="Arial", color="008000", size=10)
bld = Font(name="Arial", bold=True, size=10)
bld_bl = Font(name="Arial", bold=True, size=10, color="1B3A5C")
tf = Font(name="Arial", bold=True, size=14, color="1B3A5C")
yel = PatternFill("solid", fgColor="FFFF00")
proj = PatternFill("solid", fgColor="FFF8E1")
bdr = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"), top=Side("thin","CCCCCC"), bottom=Side("thin","CCCCCC"))
totbdr = Border(top=Side("medium","1B3A5C"), bottom=Side("double","1B3A5C"))
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%'

def hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bdr

years_h = ["", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E"]
NC = len(years_h)

wa = wb.active
wa.title = "Assumptions"
wa.sheet_properties.tabColor = "1B3A5C"
wa["A1"] = f"Key Assumptions — {TICKER} (Uber Technologies) [UPDATED POST Q4-2025]"
wa["A1"].font = tf

for i, h in enumerate(years_h):
    wa.cell(row=3, column=1+i, value=h)
hdr(wa, 3, NC)
for c in range(2, 5):
    wa.cell(row=3, column=c).fill = PatternFill("solid", fgColor="3D5A80")
for c in range(5, NC):
    wa.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E07A2F")
    wa.cell(row=3, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

assumptions_data = {
    4:  ("Revenue Growth %", [0.143, 0.177, 0.183, 0.178, 0.155, 0.136, 0.109], PCT),
    5:  ("COGS % of Revenue", [0.632, 0.625, 0.618, 0.615, 0.608, 0.600, 0.592], PCT),
    6:  ("S&M % of Revenue", [0.125, 0.115, 0.108, 0.110, 0.105, 0.100, 0.095], PCT),
    7:  ("R&D % of Revenue", [0.126, 0.120, 0.117, 0.115, 0.110, 0.105, 0.100], PCT),
    8:  ("G&A % of Revenue", [0.060, 0.055, 0.053, 0.052, 0.050, 0.048, 0.045], PCT),
    9:  ("D&A % of Revenue", [0.020, 0.020, 0.020, 0.020, 0.020, 0.020, 0.020], PCT),
    10: ("Tax Rate", [0.15, 0.18, 0.20, 0.21, 0.21, 0.21, 0.21], PCT),
    11: ("", [], ""),
    12: ("Capex % of Revenue", [0.020, 0.022, 0.022, 0.022, 0.023, 0.023, 0.023], PCT),
    13: ("AR Days", [22, 20, 19, 19, 18, 18, 17], '0'),
    14: ("AP Days", [32, 30, 28, 27, 26, 25, 25], '0'),
    15: ("Accrued Liabilities % Rev", [0.080, 0.078, 0.075, 0.073, 0.070, 0.068, 0.065], PCT),
    16: ("", [], ""),
    17: ("Interest Rate on Debt", [0.050, 0.050, 0.048, 0.048, 0.046, 0.044, 0.042], PCT),
    18: ("Shares Outstanding (M)", [2052, 2068, 2084, 2042, 2000, 1960, 1920], '#,##0'),
}

for r, (label, vals, fmt) in assumptions_data.items():
    wa.cell(row=r, column=1, value=label).font = bld if label else blk
    for i, v in enumerate(vals):
        c = 2 + i
        wa.cell(row=r, column=c, value=v).font = blue
        wa.cell(row=r, column=c).number_format = fmt
        if c >= 5:
            wa.cell(row=r, column=c).fill = yel

wa.column_dimensions["A"].width = 28
for c in range(2, NC+1):
    wa.column_dimensions[get_column_letter(c)].width = 13

# ========== INCOME STATEMENT ==========
wi = wb.create_sheet("Income Statement")
wi.sheet_properties.tabColor = "2C5F8A"
wi["A1"] = f"Income Statement — {TICKER} ($M) [UPDATED POST Q4-2025]"
wi["A1"].font = tf

for i, h in enumerate(years_h):
    wi.cell(row=3, column=1+i, value=h)
hdr(wi, 3, NC)

is_rows = [
    (4, "Revenue", True), (5, "YoY Growth %", False),
    (6, "(-) Cost of Revenue", False), (7, "Gross Profit", True), (8, "Gross Margin %", False),
    (9, "", False),
    (10, "(-) Sales & Marketing", False), (11, "(-) Research & Development", False),
    (12, "(-) General & Administrative", False),
    (13, "Total Operating Expenses", True), (14, "", False),
    (15, "Operating Income (EBIT)", True), (16, "Operating Margin %", False), (17, "", False),
    (18, "(-) Interest Expense", False), (19, "(+) Other Income / (Expense)", False),
    (20, "Pre-Tax Income", True), (21, "(-) Income Tax", False), (22, "Effective Tax Rate", False),
    (23, "Net Income", True), (24, "Net Margin %", False), (25, "", False),
    (26, "EPS (Diluted)", True), (27, "EBITDA", True), (28, "EBITDA Margin %", False),
]

for r, label, is_bold in is_rows:
    wi.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

hist_rev = [37281, 43953, 52020]
hist_cogs = [23567, 27471, 32149]
hist_sm = [4661, 5055, 5618]
hist_rd = [4698, 5274, 6086]
hist_ga = [2238, 2418, 2757]
hist_int = [465, 475, 456]
hist_other = [-600, 1800, 900]

aref = "Assumptions!"
for i in range(3):
    c = 2 + i
    cl = get_column_letter(c)
    prev_cl = get_column_letter(c-1)
    wi.cell(row=4, column=c, value=hist_rev[i]).font = blue
    wi.cell(row=5, column=c).value = f"=({cl}4-{prev_cl}4)/{prev_cl}4" if i > 0 else 0.143
    wi.cell(row=5, column=c).number_format = PCT
    wi.cell(row=6, column=c, value=hist_cogs[i]).font = blue
    wi.cell(row=7, column=c).value = f"={cl}4-{cl}6"
    wi.cell(row=8, column=c).value = f"={cl}7/{cl}4"
    wi.cell(row=8, column=c).number_format = PCT
    wi.cell(row=10, column=c, value=hist_sm[i]).font = blue
    wi.cell(row=11, column=c, value=hist_rd[i]).font = blue
    wi.cell(row=12, column=c, value=hist_ga[i]).font = blue
    wi.cell(row=13, column=c).value = f"={cl}10+{cl}11+{cl}12"
    wi.cell(row=15, column=c).value = f"={cl}7-{cl}13"
    wi.cell(row=16, column=c).value = f"={cl}15/{cl}4"
    wi.cell(row=16, column=c).number_format = PCT
    wi.cell(row=18, column=c, value=hist_int[i]).font = blue
    wi.cell(row=19, column=c, value=hist_other[i]).font = blue
    wi.cell(row=20, column=c).value = f"={cl}15-{cl}18+{cl}19"
    wi.cell(row=21, column=c).value = f"={cl}20*{aref}{cl}10"
    wi.cell(row=22, column=c).value = f"=IF({cl}20=0,0,{cl}21/{cl}20)"
    wi.cell(row=22, column=c).number_format = PCT
    wi.cell(row=23, column=c).value = f"={cl}20-{cl}21"
    wi.cell(row=24, column=c).value = f"={cl}23/{cl}4"
    wi.cell(row=24, column=c).number_format = PCT
    wi.cell(row=26, column=c).value = f"={cl}23/{aref}{cl}18"
    wi.cell(row=26, column=c).number_format = '$#,##0.00'
    wi.cell(row=27, column=c).value = f"={cl}15+{cl}4*{aref}{cl}9"
    wi.cell(row=28, column=c).value = f"={cl}27/{cl}4"
    wi.cell(row=28, column=c).number_format = PCT

for i in range(4):
    c = 5 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wi.cell(row=4, column=c).value = f"={prev}4*(1+{aref}{acl}4)"
    wi.cell(row=5, column=c).value = f"={aref}{acl}4"
    wi.cell(row=5, column=c).number_format = PCT
    wi.cell(row=6, column=c).value = f"={cl}4*{aref}{acl}5"
    wi.cell(row=7, column=c).value = f"={cl}4-{cl}6"
    wi.cell(row=8, column=c).value = f"={cl}7/{cl}4"
    wi.cell(row=8, column=c).number_format = PCT
    wi.cell(row=10, column=c).value = f"={cl}4*{aref}{acl}6"
    wi.cell(row=11, column=c).value = f"={cl}4*{aref}{acl}7"
    wi.cell(row=12, column=c).value = f"={cl}4*{aref}{acl}8"
    wi.cell(row=13, column=c).value = f"={cl}10+{cl}11+{cl}12"
    wi.cell(row=15, column=c).value = f"={cl}7-{cl}13"
    wi.cell(row=16, column=c).value = f"={cl}15/{cl}4"
    wi.cell(row=16, column=c).number_format = PCT
    wi.cell(row=18, column=c).value = f"='Balance Sheet'!{cl}14*{aref}{acl}17"
    wi.cell(row=18, column=c).font = grn
    wi.cell(row=19, column=c, value=500).font = blue
    wi.cell(row=20, column=c).value = f"={cl}15-{cl}18+{cl}19"
    wi.cell(row=21, column=c).value = f"={cl}20*{aref}{acl}10"
    wi.cell(row=22, column=c).value = f"=IF({cl}20=0,0,{cl}21/{cl}20)"
    wi.cell(row=22, column=c).number_format = PCT
    wi.cell(row=23, column=c).value = f"={cl}20-{cl}21"
    wi.cell(row=24, column=c).value = f"={cl}23/{cl}4"
    wi.cell(row=24, column=c).number_format = PCT
    wi.cell(row=26, column=c).value = f"={cl}23/{aref}{acl}18"
    wi.cell(row=26, column=c).number_format = '$#,##0.00'
    wi.cell(row=27, column=c).value = f"={cl}15+{cl}4*{aref}{acl}9"
    wi.cell(row=28, column=c).value = f"={cl}27/{cl}4"
    wi.cell(row=28, column=c).number_format = PCT

for r in [4,6,7,10,11,12,13,15,18,19,20,21,23,27]:
    for c in range(2, NC):
        wi.cell(row=r, column=c).number_format = NUM

for r in [7, 13, 15, 20, 23, 27]:
    for c in range(1, NC):
        wi.cell(row=r, column=c).font = bld_bl
        wi.cell(row=r, column=c).border = totbdr if r in [7, 15, 23] else bdr

for r in range(4, 29):
    for c in range(5, NC):
        if not wi.cell(row=r, column=c).fill or wi.cell(row=r, column=c).fill.fgColor.rgb == "00000000":
            wi.cell(row=r, column=c).fill = proj

wi.column_dimensions["A"].width = 32
for c in range(2, NC+1):
    wi.column_dimensions[get_column_letter(c)].width = 14

# ========== BALANCE SHEET ==========
wb2 = wb.create_sheet("Balance Sheet")
wb2.sheet_properties.tabColor = "3D7AB5"
wb2["A1"] = f"Balance Sheet — {TICKER} ($M) [UPDATED POST Q4-2025]"
wb2["A1"].font = tf

for i, h in enumerate(years_h):
    wb2.cell(row=3, column=1+i, value=h)
hdr(wb2, 3, NC)

bs_rows = [
    (4, "ASSETS", True), (5, "Cash & Equivalents", False), (6, "Short-Term Investments", False),
    (7, "Accounts Receivable", False), (8, "Other Current Assets", False),
    (9, "Total Current Assets", True), (10, "", False),
    (11, "PP&E (net)", False), (12, "Goodwill", False),
    (13, "Equity Investments", False), (14, "Long-Term Debt", False),
    (15, "Other Non-Current Assets", False),
    (16, "Total Assets", True), (17, "", False),
    (18, "LIABILITIES", True), (19, "Accounts Payable", False), (20, "Accrued Liabilities", False),
    (21, "Short-Term Debt", False), (22, "Other Current Liabilities", False),
    (23, "Total Current Liabilities", True), (24, "", False),
    (25, "Long-Term Debt", False), (26, "Other Non-Current Liabilities", False),
    (27, "Total Liabilities", True), (28, "", False),
    (29, "EQUITY", True), (30, "Retained Earnings", False), (31, "Other Equity", False),
    (32, "Total Equity", True), (33, "Total Liab. + Equity", True),
    (34, "Balance Check (should be 0)", False),
]

for r, label, is_bold in bs_rows:
    wb2.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

hist_bs = {
    5:  [5088, 5700, 7000],
    6:  [900, 1000, 1200],
    7:  [2249, 2403, 2708],
    8:  [1800, 2000, 2300],
    11: [2095, 2300, 2600],
    12: [16100, 16300, 16300],
    13: [13800, 15000, 15000],
    14: [9459, 9500, 9500],
    15: [3200, 3500, 3800],
    19: [1900, 2050, 2350],
    20: [2985, 3432, 3903],
    21: [500, 500, 500],
    22: [3000, 3200, 3500],
    25: [9459, 9500, 9500],
    26: [5000, 5200, 5400],
    31: [20188, 23321, 28000],
}

for r, vals in hist_bs.items():
    for i, v in enumerate(vals):
        wb2.cell(row=r, column=2+i, value=v).font = blue
        wb2.cell(row=r, column=2+i).number_format = NUM

isref = "'Income Statement'!"
for c in range(2, 5):
    cl = get_column_letter(c)
    wb2.cell(row=9, column=c).value = f"=SUM({cl}5:{cl}8)"
    wb2.cell(row=16, column=c).value = f"={cl}9+{cl}11+{cl}12+{cl}13+{cl}14+{cl}15"
    wb2.cell(row=23, column=c).value = f"=SUM({cl}19:{cl}22)"
    wb2.cell(row=27, column=c).value = f"={cl}23+{cl}25+{cl}26"
    wb2.cell(row=30, column=c).value = f"={cl}16-{cl}27-{cl}31"
    wb2.cell(row=32, column=c).value = f"={cl}30+{cl}31"
    wb2.cell(row=33, column=c).value = f"={cl}27+{cl}32"
    wb2.cell(row=34, column=c).value = f"={cl}16-{cl}33"
    wb2.cell(row=34, column=c).number_format = NUM

bsref = "'Balance Sheet'!"
for i in range(4):
    c = 5 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wb2.cell(row=5, column=c).value = f"={prev}5+'Cash Flow'!{cl}19"
    wb2.cell(row=5, column=c).font = grn
    wb2.cell(row=6, column=c).value = f"={prev}6*1.05"
    wb2.cell(row=7, column=c).value = f"={isref}{cl}4*{aref}{acl}13/365"
    wb2.cell(row=8, column=c).value = f"={prev}8*1.05"
    wb2.cell(row=9, column=c).value = f"=SUM({cl}5:{cl}8)"
    wb2.cell(row=11, column=c).value = f"={prev}11+{isref}{cl}4*{aref}{acl}12-{isref}{cl}4*{aref}{acl}9"
    wb2.cell(row=12, column=c).value = f"={prev}12"
    wb2.cell(row=13, column=c).value = f"={prev}13"
    wb2.cell(row=14, column=c).value = f"={prev}14"
    wb2.cell(row=15, column=c).value = f"={prev}15*1.03"
    wb2.cell(row=16, column=c).value = f"={cl}9+{cl}11+{cl}12+{cl}13+{cl}14+{cl}15"
    wb2.cell(row=19, column=c).value = f"={isref}{cl}6*{aref}{acl}14/365"
    wb2.cell(row=20, column=c).value = f"={isref}{cl}4*{aref}{acl}15"
    wb2.cell(row=21, column=c).value = f"={prev}21"
    wb2.cell(row=22, column=c).value = f"={prev}22*1.03"
    wb2.cell(row=23, column=c).value = f"=SUM({cl}19:{cl}22)"
    wb2.cell(row=25, column=c).value = f"={prev}25"
    wb2.cell(row=26, column=c).value = f"={prev}26*1.02"
    wb2.cell(row=27, column=c).value = f"={cl}23+{cl}25+{cl}26"
    wb2.cell(row=30, column=c).value = f"={prev}30+{isref}{cl}23"
    wb2.cell(row=31, column=c).value = f"={prev}31"
    wb2.cell(row=32, column=c).value = f"={cl}30+{cl}31"
    wb2.cell(row=33, column=c).value = f"={cl}27+{cl}32"
    wb2.cell(row=34, column=c).value = f"={cl}16-{cl}33"
    wb2.cell(row=34, column=c).number_format = NUM

for r in [5,6,7,8,9,11,12,13,14,15,16,19,20,21,22,23,25,26,27,30,31,32,33]:
    for c in range(2, NC):
        wb2.cell(row=r, column=c).number_format = NUM

for r in [9, 16, 23, 27, 32, 33]:
    for c in range(1, NC):
        wb2.cell(row=r, column=c).font = bld_bl

wb2.column_dimensions["A"].width = 30
for c in range(2, NC+1):
    wb2.column_dimensions[get_column_letter(c)].width = 14

# ========== CASH FLOW STATEMENT ==========
wc = wb.create_sheet("Cash Flow")
wc.sheet_properties.tabColor = "4A90D9"
wc["A1"] = f"Cash Flow Statement — {TICKER} ($M) [UPDATED POST Q4-2025]"
wc["A1"].font = tf

for i, h in enumerate(years_h):
    wc.cell(row=3, column=1+i, value=h)
hdr(wc, 3, NC)

cf_rows = [
    (4, "OPERATING ACTIVITIES", True), (5, "Net Income", False),
    (6, "(+) Depreciation & Amortization", False), (7, "Stock-Based Compensation", False),
    (8, "Changes in Working Capital", False),
    (9, "  Change in AR", False), (10, "  Change in AP", False),
    (11, "  Change in Accrued Liab", False), (12, "  Other Operating Changes", False),
    (13, "Cash from Operations (CFO)", True),
    (14, "", False), (15, "INVESTING ACTIVITIES", True),
    (16, "(-) Capital Expenditures", False), (17, "(-) Net Investment Changes", False),
    (18, "Cash from Investing (CFI)", True), (19, "Net Change in Cash", True),
]

for r, label, is_bold in cf_rows:
    wc.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

hist_cfo = [6867, 7490, 10100]
hist_capex = [620, 750, 900]
hist_sbc = [1900, 1800, 1700]

for i in range(3):
    c = 2 + i
    cl = get_column_letter(c)
    wc.cell(row=5, column=c).value = f"={isref}{cl}23"
    wc.cell(row=5, column=c).font = grn
    wc.cell(row=6, column=c).value = f"={isref}{cl}4*{aref}{cl}9"
    wc.cell(row=7, column=c, value=hist_sbc[i]).font = blue
    wc.cell(row=13, column=c, value=hist_cfo[i]).font = blue
    wc.cell(row=16, column=c, value=hist_capex[i]).font = blue
    wc.cell(row=17, column=c, value=-200).font = blue
    wc.cell(row=18, column=c).value = f"=-{cl}16+{cl}17"
    wc.cell(row=19, column=c).value = f"={cl}13+{cl}18"

for i in range(4):
    c = 5 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wc.cell(row=5, column=c).value = f"={isref}{cl}23"
    wc.cell(row=5, column=c).font = grn
    wc.cell(row=6, column=c).value = f"={isref}{cl}4*{aref}{acl}9"
    wc.cell(row=7, column=c, value=1500).font = blue
    wc.cell(row=9, column=c).value = f"=-({bsref}{cl}7-{bsref}{prev}7)"
    wc.cell(row=10, column=c).value = f"={bsref}{cl}19-{bsref}{prev}19"
    wc.cell(row=11, column=c).value = f"={bsref}{cl}20-{bsref}{prev}20"
    wc.cell(row=12, column=c, value=0).font = blue
    wc.cell(row=8, column=c).value = f"={cl}9+{cl}10+{cl}11+{cl}12"
    wc.cell(row=13, column=c).value = f"={cl}5+{cl}6+{cl}7+{cl}8"
    wc.cell(row=16, column=c).value = f"={isref}{cl}4*{aref}{acl}12"
    wc.cell(row=17, column=c, value=-200).font = blue
    wc.cell(row=18, column=c).value = f"=-{cl}16+{cl}17"
    wc.cell(row=19, column=c).value = f"={cl}13+{cl}18"

for r in [5,6,7,8,9,10,11,12,13,16,17,18,19]:
    for c in range(2, NC):
        wc.cell(row=r, column=c).number_format = NUM

for r in [13, 18, 19]:
    for c in range(1, NC):
        wc.cell(row=r, column=c).font = bld_bl

wc.column_dimensions["A"].width = 30
for c in range(2, NC+1):
    wc.column_dimensions[get_column_letter(c)].width = 14

out = f"coverage/{TICKER}/04-financial-model/3-statements.xlsx"
wb.save(out)
print(f"Updated 3-statement model saved to: {out}")
'''

with open(update_script, 'w') as f:
    f.write(updated_code)

result = subprocess.run(["python", update_script, TICKER], capture_output=True, text=True, cwd=os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
print(result.stdout.strip())
if result.stderr:
    print(result.stderr.strip())
