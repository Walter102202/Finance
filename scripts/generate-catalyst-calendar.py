# DISCLAIMER: Financial data hardcoded in this script is illustrative only (as of Feb 2026). Update with current data before use. Not investment advice.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import sys, datetime

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
DATE = datetime.date.today().isoformat()
wb = openpyxl.Workbook()

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
blk = Font(name="Arial", color="000000", size=10)
bld = Font(name="Arial", bold=True, size=10)
tf = Font(name="Arial", bold=True, size=14, color="1B3A5C")
bdr = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"), top=Side("thin","CCCCCC"), bottom=Side("thin","CCCCCC"))
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

def hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bdr

ws = wb.active
ws.title = "Catalyst Calendar"
ws.sheet_properties.tabColor = "1B3A5C"

ws["A1"] = f"Catalyst Calendar — {TICKER} (Uber Technologies)"
ws["A1"].font = tf
ws["A2"] = f"Last Updated: {DATE}"
ws["A2"].font = Font(name="Arial", italic=True, color="888888", size=10)

headers = ["Date", "Event", "Category", "Impact (+/-/N)", "Magnitude (H/M/L)", "Recommended Action", "Status", "Notes"]
widths = [14, 35, 16, 16, 16, 30, 14, 35]
for i, h in enumerate(headers):
    ws.cell(row=4, column=1+i, value=h)
hdr(ws, 4, len(headers))

events = [
    # UBER Earnings
    ["2026-02-04", "UBER Q4 & FY2025 Earnings", "Earnings", "+", "H", "Review results vs model; update thesis tracker", "Passed", "Rev $52B; Adj EBITDA $8.7B; FCF $9.8B. Q1 guide below consensus."],
    ["2026-05-07", "UBER Q1 2026 Earnings (Est.)", "Earnings", "TBD", "H", "Prepare earnings preview 2 weeks before; update model post", "Upcoming", "GBs guide $52-53.5B; Adj EBITDA guide $2.37-2.47B"],
    ["2026-08-06", "UBER Q2 2026 Earnings (Est.)", "Earnings", "TBD", "H", "Prepare earnings preview; key for margin trajectory", "Upcoming", "Date estimated. Watch EBITDA margin recovery."],
    ["2026-11-05", "UBER Q3 2026 Earnings (Est.)", "Earnings", "TBD", "H", "Prepare earnings preview; AV city expansion update", "Upcoming", "Date estimated based on historical pattern."],

    # AV Milestones
    ["2026-H1", "Waymo-Uber SF Launch", "AV", "+", "H", "Monitor Waymo expansion to San Francisco on Uber platform", "Upcoming", "Planned late 2026; could accelerate. Major AV catalyst."],
    ["2026-H1", "Nuro/Lucid Premium AV Service SF", "AV", "+", "M", "Track launch of Uber's own premium AV service with Lucid vehicles", "Upcoming", "$300M investment in Lucid; 20,000 vehicle commitment."],
    ["2026-Q2", "AV City Expansion Updates", "AV", "+", "H", "Monitor progress toward 15-city AV target by year-end", "Upcoming", "Currently 5 cities; need 10 more. Key execution test."],
    ["2026-H2", "Nvidia Alliance Update (100K AVs by 2027)", "AV", "+", "M", "Track progress on Nvidia partnership for 100K AVs globally", "Upcoming", "Ambitious target; signals commitment to AV fleet scale."],

    # Competitive
    ["2026-02-12", "Lyft (LYFT) Q4 2025 Earnings", "Competitive", "TBD", "M", "Monitor U.S. share trends; DashPass/Lyft+ adoption", "Passed", "Record revenue; DashPass partnership gaining traction."],
    ["2026-02-13", "DoorDash (DASH) Q4 2025 Earnings", "Competitive", "TBD", "M", "Monitor U.S. delivery share; Wolt EU performance", "Passed", "DASH growing ~38%; Wolt expanding well in Europe."],
    ["2026-05-08", "Lyft Q1 2026 Earnings (Est.)", "Competitive", "TBD", "M", "Track U.S. ride-hailing competitive dynamics", "Upcoming", ""],
    ["2026-05-09", "DoorDash Q1 2026 Earnings (Est.)", "Competitive", "TBD", "M", "Track DashPass penetration and delivery share shifts", "Upcoming", ""],
    ["2026-06-10", "Tesla Robotaxi Update (Austin)", "Competitive", "-", "M", "Monitor Tesla AV fleet size and expansion plans", "Upcoming", "44 vehicles in Austin as of early 2026. Scale risk low near-term."],

    # Regulatory
    ["2026-H1", "EU Platform Workers Directive Transposition", "Regulatory", "-", "H", "Monitor EU member state implementation deadlines; impact on driver costs", "Upcoming", "2026 transposition deadline. 20-35% driver cost increase risk."],
    ["2026-Q2", "UK Gig Worker Compensation Review", "Regulatory", "-", "M", "Track UK enforcement of worker classification ruling", "Upcoming", "Uber already paid ~GBP 600M; ongoing compliance costs."],
    ["2026-H2", "California AB5 / Prop 22 Litigation Update", "Regulatory", "N", "M", "Monitor for any changes to CA independent contractor protections", "Upcoming", "Prop 22 currently protecting IC model in CA."],

    # Macro
    ["2026-03-19", "Federal Reserve FOMC Meeting", "Macro", "TBD", "M", "Rate decision impacts consumer discretionary spending and ride-hailing demand", "Upcoming", "Consensus: hold. Watch for rate cut signals."],
    ["2026-05-07", "Federal Reserve FOMC Meeting", "Macro", "TBD", "M", "Potential rate cut could boost consumer spending and ride volumes", "Upcoming", ""],
    ["2026-06-18", "Federal Reserve FOMC Meeting", "Macro", "+", "M", "Expected first cut of 2026; positive for discretionary spending", "Upcoming", "Consensus: 25bp cut to 4.0-4.25%"],

    # Corporate
    ["2026-05-15", "UBER Annual Shareholders Meeting (Est.)", "Corporate", "N", "L", "Standard governance; watch for buyback authorization expansion", "Upcoming", "Date estimated."],
    ["2026-H2", "Potential UBER Investor Day", "Corporate", "+", "H", "Watch for updated AV strategy, medium-term margin targets, capital return framework", "Upcoming", "Not confirmed; would be significant catalyst if announced."],
]

for i, row in enumerate(events):
    r = 5 + i
    for j, val in enumerate(row):
        cell = ws.cell(row=r, column=1+j, value=val)
        cell.font = blk
        cell.alignment = lft if j in [1, 5, 7] else ctr
        cell.border = bdr

# Conditional formatting
green = PatternFill("solid", fgColor="C8E6C9")
yellow = PatternFill("solid", fgColor="FFF3CD")
gray = PatternFill("solid", fgColor="E0E0E0")
red = PatternFill("solid", fgColor="FFCDD2")

end_r = 4 + len(events)
ws.conditional_formatting.add(f"G5:G{end_r}", CellIsRule(operator="equal", formula=['"Upcoming"'], fill=yellow))
ws.conditional_formatting.add(f"G5:G{end_r}", CellIsRule(operator="equal", formula=['"Passed"'], fill=gray))
ws.conditional_formatting.add(f"G5:G{end_r}", CellIsRule(operator="equal", formula=['"Cancelled"'], fill=red))
ws.conditional_formatting.add(f"D5:D{end_r}", CellIsRule(operator="equal", formula=['"+"'], fill=green))
ws.conditional_formatting.add(f"D5:D{end_r}", CellIsRule(operator="equal", formula=['"-"'], fill=red))
ws.conditional_formatting.add(f"E5:E{end_r}", CellIsRule(operator="equal", formula=['"H"'], fill=PatternFill("solid", fgColor="FFCDD2")))
ws.conditional_formatting.add(f"E5:E{end_r}", CellIsRule(operator="equal", formula=['"M"'], fill=PatternFill("solid", fgColor="FFF3CD")))

# Summary
sr = end_r + 2
ws.cell(row=sr, column=1, value="Summary").font = Font(name="Arial", bold=True, size=11, color="2C5F8A")
ws.cell(row=sr+1, column=1, value="Total Events").font = bld
ws.cell(row=sr+1, column=2).value = f'=COUNTA(A5:A{end_r})'
ws.cell(row=sr+2, column=1, value="Upcoming").font = bld
ws.cell(row=sr+2, column=2).value = f'=COUNTIF(G5:G{end_r},"Upcoming")'
ws.cell(row=sr+3, column=1, value="Passed").font = bld
ws.cell(row=sr+3, column=2).value = f'=COUNTIF(G5:G{end_r},"Passed")'
ws.cell(row=sr+4, column=1, value="High Magnitude Upcoming").font = bld
ws.cell(row=sr+4, column=2).value = f'=COUNTIFS(E5:E{end_r},"H",G5:G{end_r},"Upcoming")'

for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

out = f"coverage/{TICKER}/07-catalyst-calendar.xlsx"
wb.save(out)
print(f"Catalyst calendar saved to: {out}")
