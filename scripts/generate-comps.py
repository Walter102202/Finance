import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

TICKER = sys.argv[1] if len(sys.argv) > 1 else "UBER"
wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="1B3A5C")
blue_font = Font(name="Arial", color="0000FF", size=10)
black_font = Font(name="Arial", color="000000", size=10)
green_font = Font(name="Arial", color="008000", size=10)
bold_font = Font(name="Arial", bold=True, size=10)
title_font = Font(name="Arial", bold=True, size=14, color="1B3A5C")
sub_font = Font(name="Arial", bold=True, size=11, color="2C5F8A")
alt_fill = PatternFill("solid", fgColor="F2F6FA")
target_fill = PatternFill("solid", fgColor="E8F5E9")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))
center = Alignment(horizontal="center", vertical="center")
left_al = Alignment(horizontal="left", vertical="center")

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

def style_data_area(ws, start_row, end_row, max_col, target_row=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = center if c > 1 else left_al
            if (r - start_row) % 2 == 1:
                cell.fill = alt_fill
            if r == target_row:
                cell.fill = target_fill

# ========== SHEET 1: Company Data ==========
ws = wb.active
ws.title = "Company Data"
ws.sheet_properties.tabColor = "1B3A5C"

ws["A1"] = f"Comparable Company Analysis — {TICKER}"
ws["A1"].font = title_font
ws["A2"] = "Source: Yahoo Finance, StockAnalysis, Company Filings (Feb 2026)"
ws["A2"].font = Font(name="Arial", italic=True, color="888888", size=9)

headers = ["Company", "Ticker", "Exchange", "Stock Price ($)", "Shares Out (M)",
           "Market Cap ($M)", "Cash ($M)", "Total Debt ($M)", "Enterprise Value ($M)",
           "Revenue TTM ($M)", "EBITDA TTM ($M)", "Net Income TTM ($M)", "Book Value ($M)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

# Companies: UBER, DASH, LYFT, GRAB, ROO (Deliveroo), TKWY (Just Eat), CART (Instacart)
companies = [
    # Company, Ticker, Exchange, Price, Shares, _, Cash, Debt, _, Revenue, EBITDA, NetInc, BookVal
    ["Uber Technologies", "UBER", "NYSE", 72.83, 2084, None, 7000, 9800, None, 52020, 6310, 10050, 28100],
    ["DoorDash", "DASH", "NASDAQ", 185, 405, None, 4600, 2500, None, 13700, 1300, 500, 7800],
    ["Lyft", "LYFT", "NASDAQ", 13.5, 395, None, 1800, 1400, None, 6300, 529, 300, 2200],
    ["Grab Holdings", "GRAB", "NASDAQ", 4.2, 3950, None, 5000, 1200, None, 3200, 490, -100, 8500],
    ["Deliveroo", "ROO", "LSE", 1.5, 1800, None, 600, 200, None, 2600, 120, -50, 1200],
    ["Just Eat Takeaway", "TKWY", "Euronext", 17, 220, None, 800, 1200, None, 3700, 300, -200, 3000],
    ["Instacart", "CART", "NASDAQ", 28, 290, None, 2100, 400, None, 3400, 850, 600, 3500],
]

for idx, comp in enumerate(companies):
    r = 5 + idx
    for c in range(1, 4):
        ws.cell(row=r, column=c, value=comp[c-1]).font = black_font
    ws.cell(row=r, column=4, value=comp[3]).font = blue_font
    ws.cell(row=r, column=5, value=comp[4]).font = blue_font
    # Market Cap = Price * Shares
    ws.cell(row=r, column=6).value = f"={get_column_letter(4)}{r}*{get_column_letter(5)}{r}"
    ws.cell(row=r, column=6).font = black_font
    ws.cell(row=r, column=6).number_format = '#,##0'
    # Cash, Debt
    ws.cell(row=r, column=7, value=comp[6]).font = blue_font
    ws.cell(row=r, column=8, value=comp[7]).font = blue_font
    # EV = Mkt Cap + Debt - Cash
    ws.cell(row=r, column=9).value = f"={get_column_letter(6)}{r}+{get_column_letter(8)}{r}-{get_column_letter(7)}{r}"
    ws.cell(row=r, column=9).font = black_font
    ws.cell(row=r, column=9).number_format = '#,##0'
    # Revenue, EBITDA, Net Inc, Book Val
    for c, val in zip([10, 11, 12, 13], [comp[9], comp[10], comp[11], comp[12]]):
        ws.cell(row=r, column=c, value=val).font = blue_font
    # Number formats
    ws.cell(row=r, column=4).number_format = '$#,##0.00'
    ws.cell(row=r, column=5).number_format = '#,##0.0'
    for c in [7, 8, 10, 11, 12, 13]:
        ws.cell(row=r, column=c).number_format = '#,##0'

style_data_area(ws, 5, 11, 13, target_row=5)

widths = [18, 10, 10, 14, 14, 16, 12, 14, 18, 16, 16, 16, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ========== SHEET 2: Trading Multiples ==========
ws2 = wb.create_sheet("Trading Multiples")
ws2.sheet_properties.tabColor = "2C5F8A"

ws2["A1"] = f"Trading Multiples — {TICKER} vs Peers"
ws2["A1"].font = title_font

mult_headers = ["Company", "Ticker", "EV/Revenue", "EV/EBITDA", "P/E", "P/Book",
                "Revenue Gr. YoY", "EBITDA Margin", "Net Margin", "ROE"]
for i, h in enumerate(mult_headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(mult_headers))

rev_growth = [0.183, 0.38, 0.14, 0.17, 0.10, -0.05, 0.15]

for idx in range(7):
    r = 4 + idx
    cd = "'Company Data'!"
    dr = 5 + idx
    ws2.cell(row=r, column=1).value = f"={cd}A{dr}"
    ws2.cell(row=r, column=1).font = green_font
    ws2.cell(row=r, column=2).value = f"={cd}B{dr}"
    ws2.cell(row=r, column=2).font = green_font
    # EV/Revenue
    ws2.cell(row=r, column=3).value = f"={cd}I{dr}/{cd}J{dr}"
    ws2.cell(row=r, column=3).font = black_font
    ws2.cell(row=r, column=3).number_format = '0.0x'
    # EV/EBITDA
    ws2.cell(row=r, column=4).value = f"={cd}I{dr}/{cd}K{dr}"
    ws2.cell(row=r, column=4).font = black_font
    ws2.cell(row=r, column=4).number_format = '0.0x'
    # P/E
    ws2.cell(row=r, column=5).value = f"={cd}F{dr}/{cd}L{dr}"
    ws2.cell(row=r, column=5).font = black_font
    ws2.cell(row=r, column=5).number_format = '0.0x'
    # P/Book
    ws2.cell(row=r, column=6).value = f"={cd}F{dr}/{cd}M{dr}"
    ws2.cell(row=r, column=6).font = black_font
    ws2.cell(row=r, column=6).number_format = '0.0x'
    # Revenue Growth
    ws2.cell(row=r, column=7, value=rev_growth[idx]).font = blue_font
    ws2.cell(row=r, column=7).number_format = '0.0%'
    # EBITDA Margin
    ws2.cell(row=r, column=8).value = f"={cd}K{dr}/{cd}J{dr}"
    ws2.cell(row=r, column=8).font = black_font
    ws2.cell(row=r, column=8).number_format = '0.0%'
    # Net Margin
    ws2.cell(row=r, column=9).value = f"={cd}L{dr}/{cd}J{dr}"
    ws2.cell(row=r, column=9).font = black_font
    ws2.cell(row=r, column=9).number_format = '0.0%'
    # ROE
    ws2.cell(row=r, column=10).value = f"={cd}L{dr}/{cd}M{dr}"
    ws2.cell(row=r, column=10).font = black_font
    ws2.cell(row=r, column=10).number_format = '0.0%'

style_data_area(ws2, 4, 10, 10, target_row=4)

# Summary stats
for label, func, r in [("Median", "MEDIAN", 12), ("Mean", "AVERAGE", 13), ("Min", "MIN", 14), ("Max", "MAX", 15)]:
    ws2.cell(row=r, column=1, value=label).font = bold_font
    for c in range(3, 11):
        col = get_column_letter(c)
        ws2.cell(row=r, column=c).value = f"={func}({col}4:{col}10)"
        ws2.cell(row=r, column=c).font = black_font
        ws2.cell(row=r, column=c).number_format = '0.0%' if c >= 7 else '0.0x'
        ws2.cell(row=r, column=c).border = thin_border
    ws2.cell(row=r, column=1).border = thin_border

# Highlight median row
for c in range(1, 11):
    ws2.cell(row=12, column=c).fill = PatternFill("solid", fgColor="FFF3CD")

# Implied valuation for UBER
ws2.cell(row=17, column=1, value=f"Implied Valuation for {TICKER}").font = sub_font
impl_headers = ["Method", "Multiple Applied", f"{TICKER} Metric ($M)", "Implied EV/Equity ($M)", "Implied Price/Share ($)"]
for i, h in enumerate(impl_headers, 1):
    ws2.cell(row=18, column=i, value=h)
style_header_row(ws2, 18, 5)

# EV/Revenue
ws2.cell(row=19, column=1, value="EV/Revenue (Median)")
ws2.cell(row=19, column=2).value = "=C12"
ws2.cell(row=19, column=2).number_format = '0.0x'
ws2.cell(row=19, column=3).value = "='Company Data'!J5"
ws2.cell(row=19, column=3).number_format = '#,##0'
ws2.cell(row=19, column=4).value = "=B19*C19"
ws2.cell(row=19, column=4).number_format = '#,##0'
ws2.cell(row=19, column=5).value = "=(D19-'Company Data'!H5+'Company Data'!G5)/'Company Data'!E5"
ws2.cell(row=19, column=5).number_format = '$#,##0'

# EV/EBITDA
ws2.cell(row=20, column=1, value="EV/EBITDA (Median)")
ws2.cell(row=20, column=2).value = "=D12"
ws2.cell(row=20, column=2).number_format = '0.0x'
ws2.cell(row=20, column=3).value = "='Company Data'!K5"
ws2.cell(row=20, column=3).number_format = '#,##0'
ws2.cell(row=20, column=4).value = "=B20*C20"
ws2.cell(row=20, column=4).number_format = '#,##0'
ws2.cell(row=20, column=5).value = "=(D20-'Company Data'!H5+'Company Data'!G5)/'Company Data'!E5"
ws2.cell(row=20, column=5).number_format = '$#,##0'

# P/E
ws2.cell(row=21, column=1, value="P/E (Median)")
ws2.cell(row=21, column=2).value = "=E12"
ws2.cell(row=21, column=2).number_format = '0.0x'
ws2.cell(row=21, column=3).value = "='Company Data'!L5"
ws2.cell(row=21, column=3).number_format = '#,##0'
ws2.cell(row=21, column=4).value = "=B21*C21"
ws2.cell(row=21, column=4).number_format = '#,##0'
ws2.cell(row=21, column=5).value = "=D21/'Company Data'!E5"
ws2.cell(row=21, column=5).number_format = '$#,##0'

for r in range(19, 22):
    for c in range(1, 6):
        ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=c).font = black_font if c > 1 else Font(name="Arial", size=10)

for i, w in enumerate([20, 10, 12, 12, 10, 10, 14, 14, 12, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ========== Save ==========
out = f"coverage/{TICKER}/03-valuation/comps-analysis.xlsx"
wb.save(out)
print(f"Comps analysis saved to: {out}")
