# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

TICKER = sys.argv[1] if len(sys.argv) > 1 else "MELI"
wb = openpyxl.Workbook()

hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1B3A5C")
blue = Font(name="Arial", color="0000FF", size=10)
blk = Font(name="Arial", color="000000", size=10)
grn = Font(name="Arial", color="008000", size=10)
bld = Font(name="Arial", bold=True, size=10)
bld_bl = Font(name="Arial", bold=True, size=10, color="1B3A5C")
tf = Font(name="Arial", bold=True, size=14, color="1B3A5C")
yel = PatternFill("solid", fgColor="FFFF00")
proj = PatternFill("solid", fgColor="FFF8E1")
bdr = Border(left=Side("thin","CCCCCC"), right=Side("thin","CCCCCC"), top=Side("thin","CCCCCC"), bottom=Side("thin","CCCCCC"))
totbdr = Border(top=Side("medium","1B3A5C"), bottom=Side("double","1B3A5C"))
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%'

def hdr(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ctr, bdr

years_h = ["", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "2029E"]
NC = len(years_h)

wa = wb.active
wa.title = "Assumptions"
wa.sheet_properties.tabColor = "1B3A5C"
wa["A1"] = f"Key Assumptions � {TICKER} (MercadoLibre) [UPDATED POST Q4-2025]"
wa["A1"].font = tf

for i, h in enumerate(years_h):
    wa.cell(row=3, column=1+i, value=h)
hdr(wa, 3, NC)
for c in range(2, 6):
    wa.cell(row=3, column=c).fill = PatternFill("solid", fgColor="3D5A80")
for c in range(6, NC):
    wa.cell(row=3, column=c).fill = PatternFill("solid", fgColor="E07A2F")
    wa.cell(row=3, column=c).font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

# Updated assumptions with 2025A actuals
assumptions_data = {
    4:  ("Revenue Growth %", [0.397, 0.374, 0.436, 0.390, 0.250, 0.200, 0.180, 0.150], PCT),
    5:  ("COGS % of Revenue", [0.522, 0.518, 0.539, 0.546, 0.540, 0.531, 0.520, 0.515], PCT),
    6:  ("S&M % of Revenue", [0.120, 0.115, 0.110, 0.112, 0.108, 0.100, 0.095, 0.090], PCT),
    7:  ("Tech & Product Dev % Rev", [0.095, 0.090, 0.088, 0.092, 0.088, 0.085, 0.082, 0.080], PCT),
    8:  ("G&A % of Revenue", [0.060, 0.055, 0.052, 0.053, 0.050, 0.048, 0.046, 0.044], PCT),
    9:  ("Provision for Doubtful % Rev", [0.045, 0.042, 0.040, 0.042, 0.038, 0.035, 0.032, 0.030], PCT),
    10: ("D&A % of Revenue", [0.030, 0.030, 0.028, 0.028, 0.028, 0.028, 0.028, 0.028], PCT),
    11: ("Tax Rate", [0.30, 0.28, 0.27, 0.285, 0.275, 0.270, 0.270, 0.270], PCT),
    12: ("", [], ""),
    13: ("Capex % of Revenue", [0.048, 0.059, 0.059, 0.043, 0.039, 0.036, 0.033, 0.031], PCT),
    14: ("AR Days", [45, 42, 40, 38, 37, 36, 35, 34], '0'),
    15: ("Inventory Days", [15, 14, 13, 12, 12, 12, 11, 11], '0'),
    16: ("AP Days", [60, 58, 55, 53, 52, 51, 50, 49], '0'),
    17: ("Accrued Liabilities % Rev", [0.08, 0.08, 0.075, 0.073, 0.070, 0.068, 0.065, 0.063], PCT),
    18: ("", [], ""),
    19: ("Interest Rate on Debt", [0.045, 0.055, 0.060, 0.058, 0.058, 0.056, 0.054, 0.052], PCT),
    20: ("Shares Outstanding (M)", [50.2, 50.4, 50.7, 51.0, 51.0, 51.0, 51.0, 51.0], '#,##0.0'),
}

for r, (label, vals, fmt) in assumptions_data.items():
    wa.cell(row=r, column=1, value=label).font = bld if label else blk
    for i, v in enumerate(vals):
        c = 2 + i
        wa.cell(row=r, column=c, value=v).font = blue
        wa.cell(row=r, column=c).number_format = fmt
        if c >= 6:
            wa.cell(row=r, column=c).fill = yel

wa.column_dimensions["A"].width = 28
for c in range(2, NC+1):
    wa.column_dimensions[get_column_letter(c)].width = 13

# ========== INCOME STATEMENT ==========
wi = wb.create_sheet("Income Statement")
wi.sheet_properties.tabColor = "2C5F8A"
wi["A1"] = f"Income Statement � {TICKER} ($M) [UPDATED POST Q4-2025]"
wi["A1"].font = tf

for i, h in enumerate(years_h):
    wi.cell(row=3, column=1+i, value=h)
hdr(wi, 3, NC)

is_rows = [
    (4, "Revenue", True), (5, "YoY Growth %", False),
    (6, "(-) Cost of Revenue", False), (7, "Gross Profit", True), (8, "Gross Margin %", False),
    (9, "", False),
    (10, "(-) Sales & Marketing", False), (11, "(-) Technology & Product Dev", False),
    (12, "(-) General & Administrative", False), (13, "(-) Provision for Doubtful Accts", False),
    (14, "Total Operating Expenses", True), (15, "", False),
    (16, "Operating Income (EBIT)", True), (17, "Operating Margin %", False), (18, "", False),
    (19, "(-) Interest Expense", False), (20, "(+) Other Income / (Expense)", False),
    (21, "Pre-Tax Income", True), (22, "(-) Income Tax", False), (23, "Effective Tax Rate", False),
    (24, "Net Income", True), (25, "Net Margin %", False), (26, "", False),
    (27, "EPS (Diluted)", True), (28, "EBITDA", True), (29, "EBITDA Margin %", False),
]

for r, label, is_bold in is_rows:
    wi.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

# Historical actuals (2022-2024 + 2025A)
hist_rev = [10537, 14473, 20777, 28900]
hist_cogs = [5500, 7500, 11200, 15780]
hist_sm = [1264, 1664, 2286, 3237]
hist_tp = [1001, 1302, 1828, 2659]
hist_ga = [632, 796, 1080, 1532]
hist_prov = [474, 608, 831, 1214]
hist_int = [58, 247, 608, 780]
hist_other = [120, 180, 250, 320]
hist_da = [316, 434, 582, 809]

aref = "Assumptions!"
for i in range(4):
    c = 2 + i
    cl = get_column_letter(c)
    prev_cl = get_column_letter(c-1)
    wi.cell(row=4, column=c, value=hist_rev[i]).font = blue
    wi.cell(row=5, column=c).value = f"=({cl}4-{prev_cl}4)/{prev_cl}4" if i > 0 else 0.397
    wi.cell(row=5, column=c).number_format = PCT
    wi.cell(row=6, column=c, value=hist_cogs[i]).font = blue
    wi.cell(row=7, column=c).value = f"={cl}4-{cl}6"
    wi.cell(row=8, column=c).value = f"={cl}7/{cl}4"
    wi.cell(row=8, column=c).number_format = PCT
    wi.cell(row=10, column=c, value=hist_sm[i]).font = blue
    wi.cell(row=11, column=c, value=hist_tp[i]).font = blue
    wi.cell(row=12, column=c, value=hist_ga[i]).font = blue
    wi.cell(row=13, column=c, value=hist_prov[i]).font = blue
    wi.cell(row=14, column=c).value = f"={cl}10+{cl}11+{cl}12+{cl}13"
    wi.cell(row=16, column=c).value = f"={cl}7-{cl}14"
    wi.cell(row=17, column=c).value = f"={cl}16/{cl}4"
    wi.cell(row=17, column=c).number_format = PCT
    wi.cell(row=19, column=c, value=hist_int[i]).font = blue
    wi.cell(row=20, column=c, value=hist_other[i]).font = blue
    wi.cell(row=21, column=c).value = f"={cl}16-{cl}19+{cl}20"
    wi.cell(row=22, column=c).value = f"={cl}21*{aref}{cl}11"
    wi.cell(row=23, column=c).value = f"=IF({cl}21=0,0,{cl}22/{cl}21)"
    wi.cell(row=23, column=c).number_format = PCT
    wi.cell(row=24, column=c).value = f"={cl}21-{cl}22"
    wi.cell(row=25, column=c).value = f"={cl}24/{cl}4"
    wi.cell(row=25, column=c).number_format = PCT
    wi.cell(row=27, column=c).value = f"={cl}24/{aref}{cl}20"
    wi.cell(row=27, column=c).number_format = '$#,##0.00'
    wi.cell(row=28, column=c).value = f"={cl}16+{cl}4*{aref}{cl}10"
    wi.cell(row=29, column=c).value = f"={cl}28/{cl}4"
    wi.cell(row=29, column=c).number_format = PCT

# Projected (2026E-2029E) - now starting from column 6
for i in range(4):
    c = 6 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wi.cell(row=4, column=c).value = f"={prev}4*(1+{aref}{acl}4)"
    wi.cell(row=5, column=c).value = f"={aref}{acl}4"
    wi.cell(row=5, column=c).number_format = PCT
    wi.cell(row=6, column=c).value = f"={cl}4*{aref}{acl}5"
    wi.cell(row=7, column=c).value = f"={cl}4-{cl}6"
    wi.cell(row=8, column=c).value = f"={cl}7/{cl}4"
    wi.cell(row=8, column=c).number_format = PCT
    wi.cell(row=10, column=c).value = f"={cl}4*{aref}{acl}6"
    wi.cell(row=11, column=c).value = f"={cl}4*{aref}{acl}7"
    wi.cell(row=12, column=c).value = f"={cl}4*{aref}{acl}8"
    wi.cell(row=13, column=c).value = f"={cl}4*{aref}{acl}9"
    wi.cell(row=14, column=c).value = f"={cl}10+{cl}11+{cl}12+{cl}13"
    wi.cell(row=16, column=c).value = f"={cl}7-{cl}14"
    wi.cell(row=17, column=c).value = f"={cl}16/{cl}4"
    wi.cell(row=17, column=c).number_format = PCT
    wi.cell(row=19, column=c).value = f"='Balance Sheet'!{cl}14*{aref}{acl}19"
    wi.cell(row=19, column=c).font = grn
    wi.cell(row=20, column=c, value=250).font = blue
    wi.cell(row=21, column=c).value = f"={cl}16-{cl}19+{cl}20"
    wi.cell(row=22, column=c).value = f"={cl}21*{aref}{acl}11"
    wi.cell(row=23, column=c).value = f"=IF({cl}21=0,0,{cl}22/{cl}21)"
    wi.cell(row=23, column=c).number_format = PCT
    wi.cell(row=24, column=c).value = f"={cl}21-{cl}22"
    wi.cell(row=25, column=c).value = f"={cl}24/{cl}4"
    wi.cell(row=25, column=c).number_format = PCT
    wi.cell(row=27, column=c).value = f"={cl}24/{aref}{acl}20"
    wi.cell(row=27, column=c).number_format = '$#,##0.00'
    wi.cell(row=28, column=c).value = f"={cl}16+{cl}4*{aref}{acl}10"
    wi.cell(row=29, column=c).value = f"={cl}28/{cl}4"
    wi.cell(row=29, column=c).number_format = PCT

for r in [4,6,7,10,11,12,13,14,16,19,20,21,22,24,28]:
    for c in range(2, NC):
        wi.cell(row=r, column=c).number_format = NUM

for r in [7, 14, 16, 21, 24, 28]:
    for c in range(1, NC):
        wi.cell(row=r, column=c).font = bld_bl
        wi.cell(row=r, column=c).border = totbdr if r in [7, 16, 24] else bdr

for r in range(4, 30):
    for c in range(6, NC):
        if not wi.cell(row=r, column=c).fill or wi.cell(row=r, column=c).fill.fgColor.rgb == "00000000":
            wi.cell(row=r, column=c).fill = proj

wi.column_dimensions["A"].width = 32
for c in range(2, NC+1):
    wi.column_dimensions[get_column_letter(c)].width = 14

# ========== BALANCE SHEET ==========
wb2 = wb.create_sheet("Balance Sheet")
wb2.sheet_properties.tabColor = "3D7AB5"
wb2["A1"] = f"Balance Sheet � {TICKER} ($M) [UPDATED POST Q4-2025]"
wb2["A1"].font = tf

for i, h in enumerate(years_h):
    wb2.cell(row=3, column=1+i, value=h)
hdr(wb2, 3, NC)

bs_rows = [
    (4, "ASSETS", True), (5, "Cash & Equivalents", False), (6, "Short-Term Investments", False),
    (7, "Accounts Receivable", False), (8, "Inventory", False), (9, "Other Current Assets", False),
    (10, "Total Current Assets", True), (11, "", False),
    (12, "PP&E (net)", False), (13, "Intangibles & Goodwill", False),
    (14, "Long-Term Debt Receivables", False), (15, "Other Non-Current Assets", False),
    (16, "Total Assets", True), (17, "", False),
    (18, "LIABILITIES", True), (19, "Accounts Payable", False), (20, "Accrued Liabilities", False),
    (21, "Short-Term Debt", False), (22, "Other Current Liabilities", False),
    (23, "Total Current Liabilities", True), (24, "", False),
    (25, "Long-Term Debt", False), (26, "Other Non-Current Liabilities", False),
    (27, "Total Liabilities", True), (28, "", False),
    (29, "EQUITY", True), (30, "Retained Earnings", False), (31, "Other Equity", False),
    (32, "Total Equity", True), (33, "Total Liab. + Equity", True),
    (34, "Balance Check (should be 0)", False),
]

for r, label, is_bold in bs_rows:
    wb2.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

# Historical BS (2022-2024 + 2025A)
hist_bs = {
    5:  [1550, 2100, 2600, 3200],
    6:  [2200, 2800, 4500, 5400],
    7:  [1300, 1665, 2277, 3010],
    8:  [433, 555, 740, 920],
    9:  [800, 1000, 1400, 1800],
    12: [1500, 2000, 3000, 3430],
    13: [350, 380, 420, 450],
    14: [4000, 5200, 7200, 12500],
    15: [1603, 1912, 3059, 3500],
    19: [1733, 2299, 3130, 4200],
    20: [843, 1158, 1558, 2110],
    21: [2500, 3200, 4500, 5200],
    22: [1200, 1500, 2000, 2400],
    25: [4500, 5200, 7200, 7800],
    26: [1133, 1184, 2457, 2700],
    31: [1827, 3071, 4351, 4351],
}

for r, vals in hist_bs.items():
    for i, v in enumerate(vals):
        wb2.cell(row=r, column=2+i, value=v).font = blue
        wb2.cell(row=r, column=2+i).number_format = NUM

isref = "'Income Statement'!"
for c in range(2, 6):
    cl = get_column_letter(c)
    wb2.cell(row=10, column=c).value = f"=SUM({cl}5:{cl}9)"
    wb2.cell(row=16, column=c).value = f"={cl}10+{cl}12+{cl}13+{cl}14+{cl}15"
    wb2.cell(row=23, column=c).value = f"=SUM({cl}19:{cl}22)"
    wb2.cell(row=27, column=c).value = f"={cl}23+{cl}25+{cl}26"
    wb2.cell(row=30, column=c).value = f"={cl}16-{cl}27-{cl}31"
    wb2.cell(row=32, column=c).value = f"={cl}30+{cl}31"
    wb2.cell(row=33, column=c).value = f"={cl}27+{cl}32"
    wb2.cell(row=34, column=c).value = f"={cl}16-{cl}33"
    wb2.cell(row=34, column=c).number_format = NUM

bsref = "'Balance Sheet'!"
for i in range(4):
    c = 6 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wb2.cell(row=5, column=c).value = f"={prev}5+'Cash Flow'!{cl}21"
    wb2.cell(row=5, column=c).font = grn
    wb2.cell(row=6, column=c).value = f"={prev}6*1.12"
    wb2.cell(row=7, column=c).value = f"={isref}{cl}4*{aref}{acl}14/365"
    wb2.cell(row=8, column=c).value = f"={isref}{cl}6*{aref}{acl}15/365"
    wb2.cell(row=9, column=c).value = f"={prev}9*1.08"
    wb2.cell(row=10, column=c).value = f"=SUM({cl}5:{cl}9)"
    wb2.cell(row=12, column=c).value = f"={prev}12+{isref}{cl}4*{aref}{acl}13-{isref}{cl}4*{aref}{acl}10"
    wb2.cell(row=13, column=c).value = f"={prev}13"
    wb2.cell(row=14, column=c).value = f"={prev}14*1.18"
    wb2.cell(row=15, column=c).value = f"={prev}15*1.06"
    wb2.cell(row=16, column=c).value = f"={cl}10+{cl}12+{cl}13+{cl}14+{cl}15"
    wb2.cell(row=19, column=c).value = f"={isref}{cl}6*{aref}{acl}16/365"
    wb2.cell(row=20, column=c).value = f"={isref}{cl}4*{aref}{acl}17"
    wb2.cell(row=21, column=c).value = f"={prev}21*1.08"
    wb2.cell(row=21, column=c).font = blue
    wb2.cell(row=22, column=c).value = f"={prev}22*1.06"
    wb2.cell(row=23, column=c).value = f"=SUM({cl}19:{cl}22)"
    wb2.cell(row=25, column=c).value = f"={prev}25*1.03"
    wb2.cell(row=25, column=c).font = blue
    wb2.cell(row=26, column=c).value = f"={prev}26*1.03"
    wb2.cell(row=27, column=c).value = f"={cl}23+{cl}25+{cl}26"
    wb2.cell(row=30, column=c).value = f"={prev}30+{isref}{cl}24"
    wb2.cell(row=31, column=c).value = f"={prev}31"
    wb2.cell(row=32, column=c).value = f"={cl}30+{cl}31"
    wb2.cell(row=33, column=c).value = f"={cl}27+{cl}32"
    wb2.cell(row=34, column=c).value = f"={cl}16-{cl}33"
    wb2.cell(row=34, column=c).number_format = NUM

for r in [5,6,7,8,9,10,12,13,14,15,16,19,20,21,22,23,25,26,27,30,31,32,33]:
    for c in range(2, NC):
        wb2.cell(row=r, column=c).number_format = NUM

for r in [10, 16, 23, 27, 32, 33]:
    for c in range(1, NC):
        wb2.cell(row=r, column=c).font = bld_bl

wb2.column_dimensions["A"].width = 30
for c in range(2, NC+1):
    wb2.column_dimensions[get_column_letter(c)].width = 14

# ========== CASH FLOW STATEMENT ==========
wc = wb.create_sheet("Cash Flow")
wc.sheet_properties.tabColor = "4A90D9"
wc["A1"] = f"Cash Flow Statement � {TICKER} ($M) [UPDATED POST Q4-2025]"
wc["A1"].font = tf

for i, h in enumerate(years_h):
    wc.cell(row=3, column=1+i, value=h)
hdr(wc, 3, NC)

cf_rows = [
    (4, "OPERATING ACTIVITIES", True), (5, "Net Income", False),
    (6, "(+) Depreciation & Amortization", False), (7, "Changes in Working Capital", False),
    (8, "  Change in AR", False), (9, "  Change in Inventory", False),
    (10, "  Change in AP", False), (11, "  Change in Accrued Liab", False),
    (12, "  Other Operating Changes", False), (13, "Cash from Operations (CFO)", True),
    (14, "", False), (15, "INVESTING ACTIVITIES", True),
    (16, "(-) Capital Expenditures", False), (17, "(-) Net Investment Changes", False),
    (18, "Cash from Investing (CFI)", True), (19, "", False),
    (20, "FINANCING ACTIVITIES", True), (21, "Net Change in Cash", True),
]

for r, label, is_bold in cf_rows:
    wc.cell(row=r, column=1, value=label).font = bld_bl if is_bold else blk

hist_cfo = [2940, 5140, 7918, 9500]
hist_cfi = [-3871, -3450, -8287, -9800]
hist_capex = [509, 860, 1221, 1230]

for i in range(4):
    c = 2 + i
    cl = get_column_letter(c)
    wc.cell(row=5, column=c).value = f"={isref}{cl}24"
    wc.cell(row=5, column=c).font = grn
    wc.cell(row=6, column=c).value = f"={isref}{cl}4*{aref}{cl}10"
    wc.cell(row=13, column=c, value=hist_cfo[i]).font = blue
    wc.cell(row=16, column=c, value=hist_capex[i]).font = blue
    cfi_net = hist_cfi[i] + hist_capex[i]
    wc.cell(row=17, column=c, value=cfi_net).font = blue
    wc.cell(row=18, column=c).value = f"=-{cl}16+{cl}17"
    wc.cell(row=21, column=c).value = f"={cl}13+{cl}18"

for i in range(4):
    c = 6 + i
    cl = get_column_letter(c)
    prev = get_column_letter(c - 1)
    acl = get_column_letter(c)
    wc.cell(row=5, column=c).value = f"={isref}{cl}24"
    wc.cell(row=5, column=c).font = grn
    wc.cell(row=6, column=c).value = f"={isref}{cl}4*{aref}{acl}10"
    wc.cell(row=8, column=c).value = f"=-({bsref}{cl}7-{bsref}{prev}7)"
    wc.cell(row=9, column=c).value = f"=-({bsref}{cl}8-{bsref}{prev}8)"
    wc.cell(row=10, column=c).value = f"={bsref}{cl}19-{bsref}{prev}19"
    wc.cell(row=11, column=c).value = f"={bsref}{cl}20-{bsref}{prev}20"
    wc.cell(row=12, column=c, value=0).font = blue
    wc.cell(row=7, column=c).value = f"={cl}8+{cl}9+{cl}10+{cl}11+{cl}12"
    wc.cell(row=13, column=c).value = f"={cl}5+{cl}6+{cl}7"
    wc.cell(row=16, column=c).value = f"={isref}{cl}4*{aref}{acl}13"
    wc.cell(row=17, column=c).value = f"=-({bsref}{cl}6-{bsref}{prev}6)-({bsref}{cl}14-{bsref}{prev}14)"
    wc.cell(row=18, column=c).value = f"=-{cl}16+{cl}17"
    wc.cell(row=21, column=c).value = f"={cl}13+{cl}18"

for r in [5,6,7,8,9,10,11,12,13,16,17,18,21]:
    for c in range(2, NC):
        wc.cell(row=r, column=c).number_format = NUM

for r in [13, 18, 21]:
    for c in range(1, NC):
        wc.cell(row=r, column=c).font = bld_bl

wc.column_dimensions["A"].width = 30
for c in range(2, NC+1):
    wc.column_dimensions[get_column_letter(c)].width = 14

out = f"coverage/{TICKER}/04-financial-model/3-statements.xlsx"
wb.save(out)
print(f"Updated 3-statement model saved to: {out}")
